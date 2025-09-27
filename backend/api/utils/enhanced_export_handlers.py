"""
Enhanced Export Handlers with Streaming Support
===============================================

This module provides optimized export handlers with streaming capabilities
for large datasets, memory management, and production-ready features.
"""

import os
import io
import gc
import csv
import json
import tempfile
import logging
from typing import Dict, List, Any, Optional, Iterator, Tuple
from datetime import datetime
from pathlib import Path

import pandas as pd
from django.conf import settings
from django.utils import timezone

# PDF Export with streaming
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter, A4
from reportlab.platypus import (
    SimpleDocTemplate, Table, TableStyle, Paragraph,
    Spacer, PageBreak, KeepTogether
)
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.pdfgen import canvas

# Excel Export with streaming
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.chart import LineChart, BarChart, Reference
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.writer.excel import ExcelWriter
from openpyxl.workbook.workbook import Workbook as OPWorkbook

logger = logging.getLogger(__name__)


class BaseStreamingExportHandler:
    """Base class for streaming export handlers"""

    def __init__(self, config: Dict[str, Any] = None):
        self.config = config or {}
        self.chunk_size = self.config.get('chunk_size', 1000)
        self.max_memory_mb = self.config.get('max_memory_mb', 512)
        self.temp_dir = self.config.get('temp_dir', tempfile.gettempdir())

        # Performance tracking
        self.total_rows_processed = 0
        self.current_memory_usage = 0

    def init_streaming_export(self, file_path: str, metadata: Dict[str, Any]):
        """Initialize streaming export - override in subclasses"""
        raise NotImplementedError

    def append_streaming_data(self, chunk_data: Dict[str, Any]):
        """Append data chunk to export - override in subclasses"""
        raise NotImplementedError

    def finalize_streaming_export(self) -> Dict[str, Any]:
        """Finalize streaming export - override in subclasses"""
        raise NotImplementedError

    def _check_memory_usage(self):
        """Check and manage memory usage"""
        import psutil
        process = psutil.Process()
        current_memory_mb = process.memory_info().rss / (1024 * 1024)

        if current_memory_mb > self.max_memory_mb:
            logger.warning(f"Memory usage high: {current_memory_mb:.1f}MB, forcing cleanup")
            gc.collect()

    def _create_temp_file(self, suffix: str) -> str:
        """Create temporary file for processing"""
        temp_file = tempfile.NamedTemporaryFile(
            suffix=suffix,
            dir=self.temp_dir,
            delete=False
        )
        temp_file.close()
        return temp_file.name


class StreamingPDFExportHandler(BaseStreamingExportHandler):
    """PDF export handler with streaming support for large datasets"""

    def __init__(self, config: Dict[str, Any] = None):
        super().__init__(config)
        self.page_size = A4 if self.config.get('page_size') == 'A4' else letter
        self.styles = getSampleStyleSheet()
        self._setup_custom_styles()

        # Streaming state
        self.pdf_file_path = None
        self.story_buffer = []
        self.current_page = 1
        self.rows_per_page = self.config.get('rows_per_page', 50)

    def _setup_custom_styles(self):
        """Set up custom paragraph styles"""
        self.title_style = ParagraphStyle(
            'CustomTitle',
            parent=self.styles['Heading1'],
            fontSize=20,
            textColor=colors.darkblue,
            spaceAfter=0.5*inch,
            alignment=1  # Center alignment
        )

        self.subtitle_style = ParagraphStyle(
            'CustomSubtitle',
            parent=self.styles['Heading2'],
            fontSize=14,
            textColor=colors.darkgrey,
            spaceAfter=0.3*inch
        )

    def init_streaming_export(self, file_path: str, metadata: Dict[str, Any]):
        """Initialize PDF streaming export"""
        self.pdf_file_path = file_path
        self.metadata = metadata

        # Create initial story with header
        self.story_buffer = []

        # Title
        title = metadata.get('title', 'Report')
        self.story_buffer.append(Paragraph(title, self.title_style))

        # Metadata
        metadata_text = self._format_metadata(metadata)
        self.story_buffer.append(Paragraph(metadata_text, self.styles['Normal']))
        self.story_buffer.append(Spacer(1, 0.3*inch))

        logger.info(f"Initialized PDF streaming export to {file_path}")

    def append_streaming_data(self, chunk_data: Dict[str, Any]):
        """Append data chunk to PDF"""
        data_chunk = chunk_data.get('data', [])
        if not data_chunk:
            return

        # Create table for this chunk
        table = self._create_data_table_chunk(data_chunk)

        # Wrap table to keep it together when possible
        wrapped_table = KeepTogether([table])
        self.story_buffer.append(wrapped_table)
        self.story_buffer.append(Spacer(1, 0.1*inch))

        self.total_rows_processed += len(data_chunk)
        self._check_memory_usage()

        logger.debug(f"Added {len(data_chunk)} rows to PDF buffer")

    def finalize_streaming_export(self) -> Dict[str, Any]:
        """Finalize PDF export"""
        try:
            # Create PDF document
            doc = SimpleDocTemplate(
                self.pdf_file_path,
                pagesize=self.page_size,
                rightMargin=0.75*inch,
                leftMargin=0.75*inch,
                topMargin=1*inch,
                bottomMargin=0.75*inch
            )

            # Build PDF
            doc.build(self.story_buffer)

            # Get file size
            file_size = os.path.getsize(self.pdf_file_path)

            logger.info(f"Completed PDF export: {self.total_rows_processed} rows, {file_size} bytes")

            return {
                'status': 'success',
                'format': 'pdf',
                'file_path': self.pdf_file_path,
                'file_size': file_size,
                'row_count': self.total_rows_processed,
                'message': f'PDF generated with {self.total_rows_processed} rows'
            }

        except Exception as e:
            logger.error(f"PDF streaming export failed: {str(e)}")
            raise

    def _create_data_table_chunk(self, data_chunk: List[Dict[str, Any]]) -> Table:
        """Create table for data chunk"""
        if not data_chunk:
            return Table([[]])

        # Get headers from first row
        headers = list(data_chunk[0].keys())

        # Prepare data for table
        table_data = [headers]

        for row in data_chunk:
            table_row = []
            for header in headers:
                value = row.get(header, '')
                # Convert to string and truncate if too long
                str_value = str(value)
                if len(str_value) > 50:
                    str_value = str_value[:47] + '...'
                table_row.append(str_value)
            table_data.append(table_row)

        # Create table
        table = Table(table_data)

        # Style the table
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('FONTSIZE', (0, 1), (-1, -1), 8),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.lightgrey]),
        ]))

        return table

    def _format_metadata(self, metadata: Dict[str, Any]) -> str:
        """Format metadata for display"""
        generated_at = metadata.get('generated_at', timezone.now().isoformat())
        user = metadata.get('user', 'system')

        return f"Generated: {generated_at} | User: {user}"


class StreamingExcelExportHandler(BaseStreamingExportHandler):
    """Excel export handler with streaming support for large datasets"""

    def __init__(self, config: Dict[str, Any] = None):
        super().__init__(config)

        # Streaming state
        self.excel_file_path = None
        self.workbook = None
        self.worksheet = None
        self.current_row = 1
        self.headers_written = False

    def init_streaming_export(self, file_path: str, metadata: Dict[str, Any]):
        """Initialize Excel streaming export"""
        self.excel_file_path = file_path
        self.metadata = metadata

        # Create workbook and worksheet
        self.workbook = Workbook()
        self.worksheet = self.workbook.active
        self.worksheet.title = "Report Data"

        # Set up styles
        self.header_font = Font(bold=True, color="FFFFFF")
        self.header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        self.border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )

        logger.info(f"Initialized Excel streaming export to {file_path}")

    def append_streaming_data(self, chunk_data: Dict[str, Any]):
        """Append data chunk to Excel worksheet"""
        data_chunk = chunk_data.get('data', [])
        if not data_chunk:
            return

        # Write headers if this is the first chunk
        if not self.headers_written:
            self._write_headers(list(data_chunk[0].keys()))

        # Write data rows
        for row_data in data_chunk:
            self.current_row += 1
            col = 1

            for header in self.headers:
                cell = self.worksheet.cell(row=self.current_row, column=col)
                value = row_data.get(header, '')

                # Handle different data types
                if isinstance(value, datetime):
                    cell.value = value.strftime('%Y-%m-%d %H:%M:%S')
                elif isinstance(value, (int, float)):
                    cell.value = value
                else:
                    cell.value = str(value)

                # Apply borders
                cell.border = self.border
                col += 1

        self.total_rows_processed += len(data_chunk)
        self._check_memory_usage()

        logger.debug(f"Added {len(data_chunk)} rows to Excel worksheet")

    def finalize_streaming_export(self) -> Dict[str, Any]:
        """Finalize Excel export"""
        try:
            # Apply formatting to the entire range
            self._apply_final_formatting()

            # Add metadata sheet
            self._add_metadata_sheet()

            # Save the workbook
            self.workbook.save(self.excel_file_path)

            # Get file size
            file_size = os.path.getsize(self.excel_file_path)

            logger.info(f"Completed Excel export: {self.total_rows_processed} rows, {file_size} bytes")

            return {
                'status': 'success',
                'format': 'excel',
                'file_path': self.excel_file_path,
                'file_size': file_size,
                'row_count': self.total_rows_processed,
                'message': f'Excel file generated with {self.total_rows_processed} rows'
            }

        except Exception as e:
            logger.error(f"Excel streaming export failed: {str(e)}")
            raise

    def _write_headers(self, headers: List[str]):
        """Write column headers to Excel worksheet"""
        self.headers = headers

        for col, header in enumerate(headers, 1):
            cell = self.worksheet.cell(row=1, column=col)
            cell.value = header
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.border = self.border
            cell.alignment = Alignment(horizontal='center')

        self.headers_written = True

    def _apply_final_formatting(self):
        """Apply final formatting to the worksheet"""
        # Auto-adjust column widths
        for column in self.worksheet.columns:
            max_length = 0
            column_letter = column[0].column_letter

            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass

            adjusted_width = min(max_length + 2, 50)  # Cap at 50 characters
            self.worksheet.column_dimensions[column_letter].width = adjusted_width

        # Freeze the header row
        self.worksheet.freeze_panes = 'A2'

    def _add_metadata_sheet(self):
        """Add metadata sheet to workbook"""
        metadata_sheet = self.workbook.create_sheet("Metadata")

        # Add metadata information
        metadata_rows = [
            ["Report Title", self.metadata.get('title', 'Report')],
            ["Generated At", self.metadata.get('generated_at', timezone.now().isoformat())],
            ["Generated By", self.metadata.get('user', 'system')],
            ["Total Rows", self.total_rows_processed],
            ["Export Format", 'Excel'],
        ]

        for row_idx, (key, value) in enumerate(metadata_rows, 1):
            metadata_sheet.cell(row=row_idx, column=1, value=key).font = Font(bold=True)
            metadata_sheet.cell(row=row_idx, column=2, value=str(value))

        # Auto-adjust column widths for metadata sheet
        metadata_sheet.column_dimensions['A'].width = 20
        metadata_sheet.column_dimensions['B'].width = 30


class StreamingCSVExportHandler(BaseStreamingExportHandler):
    """CSV export handler with streaming support for large datasets"""

    def __init__(self, config: Dict[str, Any] = None):
        super().__init__(config)

        # Streaming state
        self.csv_file_path = None
        self.csv_file = None
        self.csv_writer = None
        self.headers_written = False

    def init_streaming_export(self, file_path: str, metadata: Dict[str, Any]):
        """Initialize CSV streaming export"""
        self.csv_file_path = file_path
        self.metadata = metadata

        # Open file for writing
        self.csv_file = open(file_path, 'w', newline='', encoding='utf-8')

        logger.info(f"Initialized CSV streaming export to {file_path}")

    def append_streaming_data(self, chunk_data: Dict[str, Any]):
        """Append data chunk to CSV file"""
        data_chunk = chunk_data.get('data', [])
        if not data_chunk:
            return

        # Initialize writer and write headers if this is the first chunk
        if not self.headers_written:
            self.headers = list(data_chunk[0].keys())
            self.csv_writer = csv.DictWriter(self.csv_file, fieldnames=self.headers)
            self.csv_writer.writeheader()
            self.headers_written = True

        # Write data rows
        for row in data_chunk:
            # Convert datetime objects to strings
            processed_row = {}
            for key, value in row.items():
                if isinstance(value, datetime):
                    processed_row[key] = value.isoformat()
                else:
                    processed_row[key] = value

            self.csv_writer.writerow(processed_row)

        # Flush to ensure data is written
        self.csv_file.flush()

        self.total_rows_processed += len(data_chunk)
        self._check_memory_usage()

        logger.debug(f"Added {len(data_chunk)} rows to CSV file")

    def finalize_streaming_export(self) -> Dict[str, Any]:
        """Finalize CSV export"""
        try:
            # Close the file
            if self.csv_file:
                self.csv_file.close()

            # Get file size
            file_size = os.path.getsize(self.csv_file_path)

            logger.info(f"Completed CSV export: {self.total_rows_processed} rows, {file_size} bytes")

            return {
                'status': 'success',
                'format': 'csv',
                'file_path': self.csv_file_path,
                'file_size': file_size,
                'row_count': self.total_rows_processed,
                'message': f'CSV file generated with {self.total_rows_processed} rows'
            }

        except Exception as e:
            logger.error(f"CSV streaming export failed: {str(e)}")
            if self.csv_file:
                self.csv_file.close()
            raise


class StreamingJSONExportHandler(BaseStreamingExportHandler):
    """JSON export handler with streaming support for large datasets"""

    def __init__(self, config: Dict[str, Any] = None):
        super().__init__(config)

        # Streaming state
        self.json_file_path = None
        self.json_file = None
        self.first_chunk = True

    def init_streaming_export(self, file_path: str, metadata: Dict[str, Any]):
        """Initialize JSON streaming export"""
        self.json_file_path = file_path
        self.metadata = metadata

        # Open file for writing and start JSON array
        self.json_file = open(file_path, 'w', encoding='utf-8')
        self.json_file.write('[\n')

        logger.info(f"Initialized JSON streaming export to {file_path}")

    def append_streaming_data(self, chunk_data: Dict[str, Any]):
        """Append data chunk to JSON file"""
        data_chunk = chunk_data.get('data', [])
        if not data_chunk:
            return

        for i, row in enumerate(data_chunk):
            # Add comma if not first item
            if not self.first_chunk or i > 0:
                self.json_file.write(',\n')
            elif self.first_chunk and i == 0:
                self.first_chunk = False

            # Convert datetime objects to strings
            processed_row = {}
            for key, value in row.items():
                if isinstance(value, datetime):
                    processed_row[key] = value.isoformat()
                else:
                    processed_row[key] = value

            # Write JSON object
            json.dump(processed_row, self.json_file, indent=2, default=str)

        # Flush to ensure data is written
        self.json_file.flush()

        self.total_rows_processed += len(data_chunk)
        self._check_memory_usage()

        logger.debug(f"Added {len(data_chunk)} rows to JSON file")

    def finalize_streaming_export(self) -> Dict[str, Any]:
        """Finalize JSON export"""
        try:
            # Close JSON array and file
            if self.json_file:
                self.json_file.write('\n]')
                self.json_file.close()

            # Get file size
            file_size = os.path.getsize(self.json_file_path)

            logger.info(f"Completed JSON export: {self.total_rows_processed} rows, {file_size} bytes")

            return {
                'status': 'success',
                'format': 'json',
                'file_path': self.json_file_path,
                'file_size': file_size,
                'row_count': self.total_rows_processed,
                'message': f'JSON file generated with {self.total_rows_processed} rows'
            }

        except Exception as e:
            logger.error(f"JSON streaming export failed: {str(e)}")
            if self.json_file:
                self.json_file.close()
            raise


def get_streaming_export_handler(export_format: str, config: Dict[str, Any] = None) -> BaseStreamingExportHandler:
    """Factory function to get appropriate streaming export handler"""
    handlers = {
        'pdf': StreamingPDFExportHandler,
        'excel': StreamingExcelExportHandler,
        'xlsx': StreamingExcelExportHandler,
        'csv': StreamingCSVExportHandler,
        'json': StreamingJSONExportHandler
    }

    handler_class = handlers.get(export_format.lower())
    if not handler_class:
        raise ValueError(f"Unsupported export format: {export_format}")

    return handler_class(config)


def get_supported_streaming_formats() -> List[str]:
    """Get list of supported streaming export formats"""
    return ['pdf', 'excel', 'xlsx', 'csv', 'json']