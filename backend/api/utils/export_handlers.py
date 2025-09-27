# services/export_handlers.py
import os
import io
import pandas as pd
from typing import Dict, Any, List
from datetime import datetime
from django.conf import settings

# PDF Export Handler
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter, A4
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.graphics.shapes import Drawing
from reportlab.graphics.charts.lineplots import LinePlot
from reportlab.graphics.charts.barcharts import VerticalBarChart
from reportlab.lib.validators import Auto

# Excel Export Handler
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.chart import LineChart, BarChart, Reference
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.worksheet.dimensions import ColumnDimension


class PDFExportHandler:
    """Advanced PDF export handler with charts and professional formatting"""

    def __init__(self, config: Dict[str, Any] = None):
        self.config = config or {}
        self.page_size = A4 if self.config.get('page_size') == 'A4' else letter
        self.styles = getSampleStyleSheet()
        self._setup_custom_styles()

    def _setup_custom_styles(self):
        """Set up custom paragraph styles"""
        self.title_style = ParagraphStyle(
            'CustomTitle',
            parent=self.styles['Heading1'],
            fontSize=20,
            textColor=colors.darkblue,
            spaceAfter=0.5*inch,
            alignment=1  # Center alignment
        )

        self.subtitle_style = ParagraphStyle(
            'CustomSubtitle',
            parent=self.styles['Heading2'],
            fontSize=14,
            textColor=colors.darkgrey,
            spaceAfter=0.3*inch
        )

        self.metadata_style = ParagraphStyle(
            'MetadataStyle',
            parent=self.styles['Normal'],
            fontSize=10,
            textColor=colors.grey,
            spaceAfter=0.2*inch
        )

    def export(self, report_data: Dict[str, Any], output_path: str) -> str:
        """Export report data to PDF format with professional styling"""

        doc = SimpleDocTemplate(
            output_path,
            pagesize=self.page_size,
            rightMargin=0.75*inch,
            leftMargin=0.75*inch,
            topMargin=1*inch,
            bottomMargin=0.75*inch
        )

        story = []

        # Title and header
        title = report_data['metadata'].get('title', 'Compliance Report')
        story.append(Paragraph(title, self.title_style))

        # Report metadata
        metadata_text = self._format_metadata(report_data)
        story.append(Paragraph(metadata_text, self.metadata_style))
        story.append(Spacer(1, 0.3*inch))

        # Executive summary if available
        if 'summary' in report_data['metadata']:
            story.append(Paragraph("Executive Summary", self.subtitle_style))
            story.append(Paragraph(report_data['metadata']['summary'], self.styles['Normal']))
            story.append(Spacer(1, 0.2*inch))

        # Data table
        if report_data['data']:
            story.append(Paragraph("Report Data", self.subtitle_style))
            table = self._create_data_table(report_data['data'])
            story.append(table)
            story.append(Spacer(1, 0.3*inch))

        # Charts if configured
        if self.config.get('include_charts', True) and self._should_create_charts(report_data):
            story.append(PageBreak())
            story.append(Paragraph("Data Visualization", self.subtitle_style))
            charts = self._create_charts(report_data['data'])
            story.extend(charts)

        # Footer information
        footer_text = f"Generated on {datetime.now().strftime('%Y-%m-%d %H:%M:%S')} | Page {{page}} of {{pages}}"
        story.append(Spacer(1, 0.5*inch))
        story.append(Paragraph(footer_text, self.metadata_style))

        # Build PDF
        doc.build(story)
        return output_path

    def _format_metadata(self, report_data: Dict[str, Any]) -> str:
        """Format report metadata for display"""
        metadata = report_data['metadata']
        params = report_data.get('parameters', {})

        lines = [
            f"<b>Generated:</b> {report_data.get('generated_at', datetime.now().isoformat())}",
            f"<b>Template:</b> {metadata.get('template_name', 'Unknown')}",
            f"<b>Total Records:</b> {len(report_data['data'])}"
        ]

        if params.get('start_date') and params.get('end_date'):
            lines.append(f"<b>Date Range:</b> {params['start_date']} to {params['end_date']}")

        if params.get('venue_ids'):
            lines.append(f"<b>Venues:</b> {len(params['venue_ids'])} selected")

        return "<br/>".join(lines)

    def _create_data_table(self, data: List[Dict[str, Any]]) -> Table:
        """Create a formatted data table"""
        if not data:
            return Table([["No data available"]])

        # Prepare table data
        headers = list(data[0].keys())
        table_data = [headers]

        for row in data:
            table_row = []
            for header in headers:
                value = row.get(header, '')
                if isinstance(value, float):
                    value = f"{value:.2f}"
                elif isinstance(value, datetime):
                    value = value.strftime('%Y-%m-%d %H:%M')
                elif value is None:
                    value = ''
                table_row.append(str(value))
            table_data.append(table_row)

        # Create table
        table = Table(table_data, repeatRows=1)

        # Apply styling
        table.setStyle(TableStyle([
            # Header styling
            ('BACKGROUND', (0, 0), (-1, 0), colors.navy),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),

            # Data styling
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('FONTSIZE', (0, 1), (-1, -1), 9),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),

            # Alternating row colors
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.lightgrey]),
        ]))

        return table

    def _should_create_charts(self, report_data: Dict[str, Any]) -> bool:
        """Determine if charts should be created"""
        data = report_data['data']
        if not data or len(data) < 2:
            return False

        # Check for numeric columns suitable for charting
        numeric_columns = []
        for key, value in data[0].items():
            if isinstance(value, (int, float)) and key not in ['id', 'venue_id']:
                numeric_columns.append(key)

        return len(numeric_columns) > 0

    def _create_charts(self, data: List[Dict[str, Any]]) -> List:
        """Create charts based on data content"""
        charts = []

        # Example: Compliance rate trend chart
        if self._has_numeric_column(data, 'compliance_rate'):
            chart_drawing = Drawing(400, 200)
            chart = LinePlot()
            chart.x = 50
            chart.y = 50
            chart.height = 125
            chart.width = 300

            # Prepare data for chart
            chart_data = [(i, row.get('compliance_rate', 0)) for i, row in enumerate(data[:10])]
            chart.data = [chart_data]
            chart.lines[0].strokeColor = colors.blue
            chart.lines[0].strokeWidth = 2

            chart_drawing.add(chart)
            charts.append(chart_drawing)
            charts.append(Spacer(1, 0.2*inch))

        return charts

    def _has_numeric_column(self, data: List[Dict[str, Any]], column: str) -> bool:
        """Check if data has a specific numeric column"""
        if not data:
            return False
        return column in data[0] and isinstance(data[0][column], (int, float))


class ExcelExportHandler:
    """Advanced Excel export handler with multiple sheets, charts, and formatting"""

    def __init__(self, config: Dict[str, Any] = None):
        self.config = config or {}
        self.workbook = None

    def export(self, report_data: Dict[str, Any], output_path: str) -> str:
        """Export report data to Excel format with professional formatting"""

        self.workbook = Workbook()

        # Remove default sheet
        self.workbook.remove(self.workbook.active)

        # Create main data sheet
        data_sheet = self.workbook.create_sheet("Report Data")
        self._populate_data_sheet(data_sheet, report_data)

        # Create summary sheet
        if self._should_create_summary(report_data):
            summary_sheet = self.workbook.create_sheet("Summary", 0)  # Insert at beginning
            self._populate_summary_sheet(summary_sheet, report_data)

        # Create charts sheet if configured
        if self.config.get('include_charts', True) and len(report_data['data']) > 1:
            charts_sheet = self.workbook.create_sheet("Charts")
            self._add_charts(charts_sheet, data_sheet, report_data)

        # Apply workbook-level formatting
        self._apply_workbook_formatting()

        # Save workbook
        self.workbook.save(output_path)
        return output_path

    def _populate_data_sheet(self, sheet, report_data: Dict[str, Any]):
        """Populate data sheet with formatted report data"""

        # Title and metadata
        sheet['A1'] = report_data['metadata'].get('title', 'Compliance Report')
        sheet['A1'].font = Font(size=18, bold=True, color="1f4e79")
        sheet.merge_cells('A1:E1')

        # Metadata section
        metadata_row = 3
        sheet[f'A{metadata_row}'] = f"Generated: {report_data.get('generated_at', datetime.now().isoformat())}"
        sheet[f'A{metadata_row + 1}'] = f"Records: {len(report_data['data'])}"

        # Add parameters if available
        params = report_data.get('parameters', {})
        if params.get('start_date') and params.get('end_date'):
            sheet[f'A{metadata_row + 2}'] = f"Date Range: {params['start_date']} to {params['end_date']}"
            metadata_row += 1

        # Data section starts after metadata
        data_start_row = metadata_row + 3

        # Convert to DataFrame for easier handling
        if report_data['data']:
            df = pd.DataFrame(report_data['data'])

            # Add DataFrame to sheet starting from data_start_row
            for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True)):
                for c_idx, value in enumerate(row):
                    cell = sheet.cell(row=data_start_row + r_idx, column=c_idx + 1, value=value)

                    # Format header row
                    if r_idx == 0:
                        cell.font = Font(bold=True, color="ffffff")
                        cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                        cell.alignment = Alignment(horizontal="center", vertical="center")
                        cell.border = Border(
                            left=Side(style='thin'),
                            right=Side(style='thin'),
                            top=Side(style='thin'),
                            bottom=Side(style='thin')
                        )
                    else:
                        # Format data rows
                        cell.alignment = Alignment(horizontal="left", vertical="center")
                        if r_idx % 2 == 0:  # Alternating row colors
                            cell.fill = PatternFill(start_color="f2f2f2", end_color="f2f2f2", fill_type="solid")

        # Auto-adjust column widths
        self._auto_adjust_columns(sheet)

    def _populate_summary_sheet(self, sheet, report_data: Dict[str, Any]):
        """Create a summary sheet with key metrics"""

        sheet['A1'] = "Executive Summary"
        sheet['A1'].font = Font(size=16, bold=True, color="1f4e79")

        # Calculate summary statistics
        data = report_data['data']
        if not data:
            sheet['A3'] = "No data available for summary"
            return

        df = pd.DataFrame(data)

        # Basic statistics
        row = 3
        sheet[f'A{row}'] = "Total Records:"
        sheet[f'B{row}'] = len(data)
        row += 1

        # Numeric column summaries
        numeric_columns = df.select_dtypes(include=['number']).columns
        for col in numeric_columns:
            if col not in ['id', 'venue_id']:  # Skip ID columns
                sheet[f'A{row}'] = f"Average {col.replace('_', ' ').title()}:"
                sheet[f'B{row}'] = round(df[col].mean(), 2)
                row += 1

                sheet[f'A{row}'] = f"Total {col.replace('_', ' ').title()}:"
                sheet[f'B{row}'] = round(df[col].sum(), 2)
                row += 1

        # Apply formatting to summary data
        for i in range(3, row):
            sheet[f'A{i}'].font = Font(bold=True)
            sheet[f'B{i}'].alignment = Alignment(horizontal="right")

    def _add_charts(self, charts_sheet, data_sheet, report_data: Dict[str, Any]):
        """Add charts to the charts sheet"""

        data = report_data['data']
        if not data or len(data) < 2:
            charts_sheet['A1'] = "Insufficient data for charts"
            return

        df = pd.DataFrame(data)
        numeric_columns = df.select_dtypes(include=['number']).columns

        chart_row = 1

        # Create bar chart for numeric data
        if len(numeric_columns) > 0:
            chart = BarChart()
            chart.title = f"{numeric_columns[0].replace('_', ' ').title()} Distribution"
            chart.x_axis.title = 'Records'
            chart.y_axis.title = numeric_columns[0].replace('_', ' ').title()

            # Reference data from data sheet (assuming data starts at row 6)
            data_start_row = 6  # Adjust based on your data layout
            data_range = Reference(data_sheet,
                                 min_col=list(df.columns).index(numeric_columns[0]) + 1,
                                 min_row=data_start_row + 1,
                                 max_row=data_start_row + min(len(data), 20))  # Limit to 20 records

            chart.add_data(data_range)
            charts_sheet.add_chart(chart, f"A{chart_row}")
            chart_row += 15

        # Add line chart for trend data if applicable
        if 'compliance_rate' in df.columns:
            line_chart = LineChart()
            line_chart.title = "Compliance Rate Trend"
            line_chart.x_axis.title = 'Period'
            line_chart.y_axis.title = 'Compliance Rate (%)'

            # Reference compliance rate data
            if 'compliance_rate' in df.columns:
                rate_col_idx = list(df.columns).index('compliance_rate') + 1
                rate_range = Reference(data_sheet,
                                     min_col=rate_col_idx,
                                     min_row=data_start_row + 1,
                                     max_row=data_start_row + min(len(data), 20))

                line_chart.add_data(rate_range)
                charts_sheet.add_chart(line_chart, f"A{chart_row}")

    def _should_create_summary(self, report_data: Dict[str, Any]) -> bool:
        """Determine if a summary sheet should be created"""
        return len(report_data['data']) > 5  # Create summary for datasets with more than 5 records

    def _auto_adjust_columns(self, sheet):
        """Auto-adjust column widths based on content"""
        from openpyxl.utils import get_column_letter

        for col_num in range(1, sheet.max_column + 1):
            max_length = 0
            column_letter = get_column_letter(col_num)

            for row_num in range(1, sheet.max_row + 1):
                cell = sheet.cell(row=row_num, column=col_num)
                try:
                    if cell.value and len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass

            adjusted_width = min(max_length + 2, 50)  # Cap at 50 characters
            sheet.column_dimensions[column_letter].width = adjusted_width

    def _apply_workbook_formatting(self):
        """Apply consistent formatting across the workbook"""
        for sheet in self.workbook.worksheets:
            # Set default font for the entire sheet
            for row in sheet.iter_rows():
                for cell in row:
                    if cell.font == Font():  # Default font
                        cell.font = Font(name='Calibri', size=11)


# Export format registry
EXPORT_HANDLERS = {
    'pdf': PDFExportHandler,
    'excel': ExcelExportHandler,
    'xlsx': ExcelExportHandler,  # Alias for excel
}


def get_export_handler(format_type: str, config: Dict[str, Any] = None):
    """Get the appropriate export handler for a given format"""
    handler_class = EXPORT_HANDLERS.get(format_type.lower())
    if not handler_class:
        raise ValueError(f"Unsupported export format: {format_type}")

    return handler_class(config)


def get_available_formats() -> Dict[str, Dict[str, Any]]:
    """Get information about all available export formats"""
    return {
        'pdf': {
            'name': 'PDF Document',
            'mime_type': 'application/pdf',
            'supports_charts': True,
            'supports_formatting': True,
            'max_size_mb': 50,
            'description': 'Professional PDF reports with charts and styling'
        },
        'excel': {
            'name': 'Excel Workbook',
            'mime_type': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            'supports_charts': True,
            'supports_multiple_sheets': True,
            'max_size_mb': 100,
            'description': 'Multi-sheet Excel workbooks with charts and formatting'
        },
        'csv': {
            'name': 'CSV Data',
            'mime_type': 'text/csv',
            'supports_charts': False,
            'supports_formatting': False,
            'max_size_mb': 200,
            'description': 'Simple comma-separated values format'
        },
        'json': {
            'name': 'JSON Data',
            'mime_type': 'application/json',
            'supports_nested_data': True,
            'max_size_mb': 100,
            'description': 'Structured JSON data format'
        }
    }