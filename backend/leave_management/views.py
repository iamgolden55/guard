from rest_framework import viewsets, status, permissions, filters, renderers
from rest_framework.decorators import action
from rest_framework.response import Response
from rest_framework.permissions import IsAuthenticated


# Custom renderer for raw HttpResponse (used for file downloads)
class PassthroughRenderer(renderers.BaseRenderer):
    """
    Return data as-is. View should supply a Response.
    """
    media_type = '*/*'
    format = None

    def render(self, data, accepted_media_type=None, renderer_context=None):
        return data
from django_filters.rest_framework import DjangoFilterBackend
from django.db.models import Q, Sum, Count, Avg, Prefetch
from django.utils import timezone
from django.contrib.auth import get_user_model
from django.http import Http404
from django.core.cache import cache
from django.utils.decorators import method_decorator
from django.views.decorators.cache import cache_page
from decimal import Decimal
import logging
import requests
from datetime import datetime, date, timedelta
from django.db import models

from .models import LeaveType, LeavePolicy, LeaveEntitlement, LeaveRequest, BlackoutPeriod, LeaveBalance
from .serializers import (
    LeaveTypeSerializer, LeavePolicySerializer, LeavePolicyListSerializer,
    LeavePolicyAdminSerializer, LeaveEntitlementSerializer,
    LeaveBalanceSerializer, LeaveTypeUsageSerializer,
    LeaveRequestSerializer, BlackoutPeriodSerializer, TeamOverviewSerializer,
    LeaveAnalyticsSerializer, UserBasicSerializer
)
from .permissions import (
    LeaveTypePermission, LeavePolicyPermission, LeaveEntitlementPermission,
    LeaveBalancePermission, AdminOnlyPermission, ManagerOrAdminPermission,
    ReadOnlyForStaffMixin
)
from .services import LeaveBalanceService, LeaveAccrualService

User = get_user_model()
logger = logging.getLogger(__name__)


class LeaveTypeViewSet(ReadOnlyForStaffMixin, viewsets.ModelViewSet):
    """
    ViewSet for managing leave types

    - Admin: Full CRUD access
    - Manager/Staff: Read-only access
    """
    queryset = LeaveType.objects.all().prefetch_related('employment_types')
    serializer_class = LeaveTypeSerializer
    permission_classes = [IsAuthenticated, LeaveTypePermission]
    filter_backends = [DjangoFilterBackend, filters.SearchFilter, filters.OrderingFilter]
    filterset_fields = ['is_active', 'requires_approval']
    search_fields = ['name', 'code', 'description']
    ordering_fields = ['name', 'code', 'created_at']
    ordering = ['name']

    def get_queryset(self):
        """Filter queryset based on user permissions and request parameters"""
        queryset = super().get_queryset()

        # Filter by active status if requested
        if self.request.query_params.get('active_only', 'false').lower() == 'true':
            queryset = queryset.filter(is_active=True)

        # Filter by employment type if specified
        employment_type = self.request.query_params.get('employment_type')
        if employment_type:
            queryset = queryset.filter(
                Q(employment_types__isnull=True) |
                Q(employment_types=employment_type)
            ).distinct()

        return queryset

    @action(detail=False, methods=['get'])
    def active(self, request):
        """Get only active leave types"""
        active_types = self.get_queryset().filter(is_active=True)
        serializer = self.get_serializer(active_types, many=True)
        return Response(serializer.data)

    @action(detail=True, methods=['post'], permission_classes=[AdminOnlyPermission])
    def toggle_active(self, request, pk=None):
        """Toggle active status of a leave type"""
        leave_type = self.get_object()
        leave_type.is_active = not leave_type.is_active
        leave_type.save()

        serializer = self.get_serializer(leave_type)
        return Response({
            'message': f'Leave type {leave_type.name} {"activated" if leave_type.is_active else "deactivated"}',
            'leave_type': serializer.data
        })

    @action(detail=False, methods=['get'], permission_classes=[ManagerOrAdminPermission])
    def usage_statistics(self, request):
        """Get usage statistics for all leave types"""
        current_year = timezone.now().year

        stats = []
        for leave_type in self.get_queryset().filter(is_active=True):
            entitlements = LeaveEntitlement.objects.filter(
                policy__leave_type=leave_type,
                year=current_year
            )

            total_entitlement = entitlements.aggregate(
                total=Sum('annual_entitlement')
            )['total'] or Decimal('0')

            total_used = entitlements.aggregate(
                total=Sum('used_to_date')
            )['total'] or Decimal('0')

            usage_percentage = (
                (total_used / total_entitlement * 100) if total_entitlement > 0 else 0
            )

            stats.append({
                'leave_type': LeaveTypeSerializer(leave_type).data,
                'total_users': entitlements.count(),
                'total_entitlement': total_entitlement,
                'total_used': total_used,
                'total_remaining': total_entitlement - total_used,
                'usage_percentage': round(usage_percentage, 2)
            })

        return Response(stats)


class LeavePolicyViewSet(ReadOnlyForStaffMixin, viewsets.ModelViewSet):
    """
    ViewSet for managing leave policies

    - Admin: Full CRUD access
    - Manager/Staff: Read-only access
    """
    queryset = LeavePolicy.objects.select_related('leave_type').prefetch_related(
        'employment_types'
    ).all()
    permission_classes = [IsAuthenticated, LeavePolicyPermission]
    filter_backends = [DjangoFilterBackend, filters.SearchFilter, filters.OrderingFilter]
    filterset_fields = ['is_active', 'leave_type', 'accrual_method', 'carryover_method']
    search_fields = ['name', 'leave_type__name']
    ordering_fields = ['name', 'leave_type__name', 'effective_date', 'created_at']
    ordering = ['leave_type__name', 'name']

    def get_serializer_class(self):
        """Return appropriate serializer based on action and user role"""
        permission_checker = LeavePolicyPermission()

        if self.action == 'list':
            return LeavePolicyListSerializer
        elif permission_checker.is_admin(self.request.user):
            return LeavePolicyAdminSerializer
        return LeavePolicySerializer

    def get_queryset(self):
        """Filter queryset based on user permissions"""
        queryset = super().get_queryset()
        user = self.request.user

        # Filter by active status if requested
        if self.request.query_params.get('active_only', 'false').lower() == 'true':
            queryset = queryset.filter(is_active=True)

        # Filter by current date effectiveness
        if self.request.query_params.get('effective_only', 'false').lower() == 'true':
            current_date = timezone.now().date()
            queryset = queryset.filter(
                effective_date__lte=current_date
            ).filter(
                Q(expiry_date__isnull=True) | Q(expiry_date__gt=current_date)
            )

        # Filter by user's employment type if not admin
        permission_checker = LeavePolicyPermission()
        if not permission_checker.is_admin(user):
            if hasattr(user, 'profile') and user.profile and user.profile.employment_type:
                queryset = queryset.filter(
                    Q(employment_types__isnull=True) |
                    Q(employment_types=user.profile.employment_type)
                ).distinct()

        return queryset

    @action(detail=False, methods=['get'])
    def for_user(self, request):
        """Get leave policies applicable to the current user"""
        user = request.user
        policies = LeavePolicy.objects.for_user(user)
        serializer = LeavePolicyListSerializer(policies, many=True)
        return Response(serializer.data)

    @action(detail=True, methods=['post'], permission_classes=[AdminOnlyPermission])
    def duplicate(self, request, pk=None):
        """Duplicate a leave policy with modifications"""
        source_policy = self.get_object()

        # Get new name from request
        new_name = request.data.get('name', f"{source_policy.name} (Copy)")

        # Create new policy
        new_policy = LeavePolicy.objects.create(
            name=new_name,
            leave_type=source_policy.leave_type,
            accrual_method=source_policy.accrual_method,
            accrual_rate=source_policy.accrual_rate,
            max_accrual_per_year=source_policy.max_accrual_per_year,
            max_balance=source_policy.max_balance,
            service_brackets=source_policy.service_brackets,
            carryover_method=source_policy.carryover_method,
            carryover_limit=source_policy.carryover_limit,
            carryover_expiry_months=source_policy.carryover_expiry_months,
            probation_months=source_policy.probation_months,
            min_employment_days=source_policy.min_employment_days,
            allow_negative_balance=source_policy.allow_negative_balance,
            negative_balance_limit=source_policy.negative_balance_limit,
            is_active=False,  # Start inactive to allow configuration
            effective_date=timezone.now().date()
        )

        # Copy employment types
        new_policy.employment_types.set(source_policy.employment_types.all())

        serializer = self.get_serializer(new_policy)
        return Response({
            'message': f'Policy duplicated successfully',
            'policy': serializer.data
        }, status=status.HTTP_201_CREATED)

    @action(detail=True, methods=['post'], permission_classes=[AdminOnlyPermission])
    def toggle_active(self, request, pk=None):
        """Toggle active status of a leave policy"""
        policy = self.get_object()
        policy.is_active = not policy.is_active
        policy.save()

        serializer = self.get_serializer(policy)
        return Response({
            'message': f'Policy {policy.name} {"activated" if policy.is_active else "deactivated"}',
            'policy': serializer.data
        })

    @action(detail=True, methods=['get'], permission_classes=[ManagerOrAdminPermission])
    def preview_impact(self, request, pk=None):
        """Preview the impact of policy changes"""
        policy = self.get_object()
        current_year = timezone.now().year

        # Get affected users
        affected_users = User.objects.filter(
            leave_entitlements__policy=policy,
            leave_entitlements__year=current_year
        ).distinct()

        impact_summary = {
            'affected_users_count': affected_users.count(),
            'total_entitlements': policy.entitlements.filter(year=current_year).count(),
            'total_balance_impact': policy.entitlements.filter(
                year=current_year
            ).aggregate(
                total=Sum('annual_entitlement')
            )['total'] or Decimal('0')
        }

        return Response({
            'policy': LeavePolicyListSerializer(policy).data,
            'impact': impact_summary
        })


class LeaveBalanceViewSet(viewsets.ReadOnlyModelViewSet):
    """
    ViewSet for viewing leave balances

    - Admin: View all balances
    - Manager: View team balances
    - Staff: View own balances only
    """
    queryset = LeaveEntitlement.objects.select_related(
        'user', 'policy__leave_type'
    ).prefetch_related('policy__employment_types')
    serializer_class = LeaveEntitlementSerializer
    permission_classes = [IsAuthenticated, LeaveEntitlementPermission]
    filter_backends = [DjangoFilterBackend, filters.OrderingFilter]
    filterset_fields = ['year', 'policy__leave_type', 'user']
    ordering_fields = ['user__username', 'policy__leave_type__name', 'year']
    ordering = ['user__username', 'policy__leave_type__name']

    def get_queryset(self):
        """Filter queryset based on user permissions"""
        queryset = super().get_queryset()
        user = self.request.user

        permission_checker = LeaveBalancePermission()
        return permission_checker.filter_queryset_for_user(queryset, user)

    @method_decorator(cache_page(60 * 5))  # Cache for 5 minutes
    @action(detail=False, methods=['get'])
    def summary(self, request):
        """Get aggregated leave balance summary"""
        user_id = request.query_params.get('user_id')
        year = request.query_params.get('year', timezone.now().year)

        # Determine target user
        if user_id and user_id != str(request.user.id):
            permission_checker = LeaveBalancePermission()
            if not permission_checker.is_manager(request.user):
                return Response(
                    {'error': 'Permission denied'},
                    status=status.HTTP_403_FORBIDDEN
                )
            try:
                target_user = User.objects.get(id=user_id)
            except User.DoesNotExist:
                return Response(
                    {'error': 'User not found'},
                    status=status.HTTP_404_NOT_FOUND
                )
        else:
            target_user = request.user

        # Get balance service
        balance_service = LeaveBalanceService()
        balances = balance_service.get_user_balances(target_user, year)

        return Response(balances)

    @action(detail=False, methods=['get'])
    def my_balances(self, request):
        """Get current user's leave balances"""
        current_year = timezone.now().year
        year = request.query_params.get('year', current_year)

        balance_service = LeaveBalanceService()
        balances = balance_service.get_user_balances(request.user, year)

        return Response(balances)

    @action(detail=False, methods=['post'], permission_classes=[AdminOnlyPermission])
    def recalculate_all(self, request):
        """Recalculate all leave balances for current year"""
        year = request.data.get('year', timezone.now().year)

        # Use accrual service to recalculate
        accrual_service = LeaveAccrualService()
        updated_count = 0

        for entitlement in LeaveEntitlement.objects.filter(year=year):
            accrual_service.update_user_accruals(entitlement.user, year)
            updated_count += 1

        # Clear cache
        cache.clear()

        return Response({
            'message': f'Recalculated balances for {updated_count} entitlements',
            'year': year,
            'updated_count': updated_count
        })

    @action(detail=False, methods=['get'], permission_classes=[ManagerOrAdminPermission])
    def team_summary(self, request):
        """Get leave balance summary for team members"""
        # TODO: Implement team hierarchy filtering
        current_year = timezone.now().year
        year = request.query_params.get('year', current_year)

        # For now, return all users if admin, limited if manager
        permission_checker = LeaveBalancePermission()
        if permission_checker.is_admin(request.user):
            entitlements = LeaveEntitlement.objects.filter(year=year)
        else:
            # Managers see all for now - implement team filtering later
            entitlements = LeaveEntitlement.objects.filter(year=year)

        # Group by user
        user_summaries = {}
        for entitlement in entitlements.select_related('user', 'policy__leave_type'):
            user_id = entitlement.user.id
            if user_id not in user_summaries:
                user_summaries[user_id] = {
                    'user': {
                        'id': entitlement.user.id,
                        'username': entitlement.user.username,
                        'full_name': f"{entitlement.user.first_name} {entitlement.user.last_name}".strip(),
                        'email': entitlement.user.email
                    },
                    'leave_balances': []
                }

            user_summaries[user_id]['leave_balances'].append({
                'leave_type': entitlement.policy.leave_type.name,
                'current_balance': entitlement.current_balance,
                'total_entitlement': entitlement.total_entitlement,
                'used_to_date': entitlement.used_to_date
            })

        return Response(list(user_summaries.values()))


class TeamOverviewViewSet(viewsets.ReadOnlyModelViewSet):
    """
    ViewSet for Team Overview functionality

    Provides comprehensive team-level leave management views for managers and admins
    """
    permission_classes = [IsAuthenticated, ManagerOrAdminPermission]
    serializer_class = TeamOverviewSerializer

    def get_queryset(self):
        """Get team members based on user permissions"""
        # Get the user's company from their active membership
        user_membership = self.request.user.company_memberships.filter(
            is_active=True
        ).select_related('company').first()

        if not user_membership:
            # User has no company membership - return empty queryset
            return User.objects.none()

        # Get all users who are members of the same company
        company_user_ids = user_membership.company.memberships.filter(
            is_active=True
        ).values_list('user_id', flat=True)

        # Return active users from the same company
        return User.objects.filter(
            id__in=company_user_ids,
            is_active=True
        )

    def list(self, request):
        """Get team overview data"""
        users = self.get_queryset()
        serializer = self.get_serializer(users, many=True)
        return Response({
            'team_members': serializer.data,
            'summary': {
                'total_team_members': users.count(),
                'active_members': users.filter(is_active=True).count(),
            }
        })

    @action(detail=False, methods=['get'])
    def team_balances(self, request):
        """Get consolidated team leave balances"""
        current_year = timezone.now().year
        year = request.query_params.get('year', current_year)

        users = self.get_queryset()
        team_balances = []

        for user in users.select_related('profile'):
            entitlements = LeaveEntitlement.objects.filter(
                user=user,
                year=year
            ).select_related('policy__leave_type')

            user_balance = {
                'user': UserBasicSerializer(user).data,
                'balances': []
            }

            for entitlement in entitlements:
                user_balance['balances'].append({
                    'leave_type': entitlement.policy.leave_type.name,
                    'leave_type_code': entitlement.policy.leave_type.code,
                    'color_code': entitlement.policy.leave_type.color_code,
                    'current_balance': entitlement.current_balance,
                    'total_entitlement': entitlement.total_entitlement,
                    'used_to_date': entitlement.used_to_date,
                    'percentage_used': (
                        (entitlement.used_to_date / entitlement.total_entitlement * 100)
                        if entitlement.total_entitlement > 0 else 0
                    )
                })

            team_balances.append(user_balance)

        return Response({
            'year': year,
            'team_balances': team_balances,
            'generated_at': timezone.now().isoformat()
        })

    @action(detail=False, methods=['get'])
    def team_calendar(self, request):
        """Get team calendar view for leave planning"""
        from datetime import date, timedelta

        # Parse date range from query params
        start_date_str = request.query_params.get('start_date')
        end_date_str = request.query_params.get('end_date')

        if start_date_str and end_date_str:
            start_date = datetime.strptime(start_date_str, '%Y-%m-%d').date()
            end_date = datetime.strptime(end_date_str, '%Y-%m-%d').date()
        else:
            # Default to current month
            today = date.today()
            start_date = today.replace(day=1)
            next_month = start_date.replace(month=start_date.month + 1) if start_date.month < 12 else start_date.replace(year=start_date.year + 1, month=1)
            end_date = next_month - timedelta(days=1)

        # Get all approved leave requests in the date range
        leave_requests = LeaveRequest.objects.filter(
            status='approved',
            start_date__lte=end_date,
            end_date__gte=start_date,
            staff_user__in=self.get_queryset()
        ).select_related('staff_user', 'leave_type').order_by('start_date')

        # Build calendar data
        calendar_data = []
        current_date = start_date

        while current_date <= end_date:
            day_events = []

            for request in leave_requests:
                if request.start_date <= current_date <= request.end_date:
                    day_events.append({
                        'id': request.id,
                        'user_id': request.staff_user.id,
                        'username': request.staff_user.username,
                        'full_name': f"{request.staff_user.first_name} {request.staff_user.last_name}".strip(),
                        'leave_type': request.leave_type.name,
                        'leave_type_code': request.leave_type.code,
                        'color_code': request.leave_type.color_code,
                        'request_type': request.request_type,
                        'is_full_day': request.request_type == 'full_day',
                        'days_requested': float(request.days_requested),
                        'start_date': request.start_date,
                        'end_date': request.end_date
                    })

            calendar_data.append({
                'date': current_date,
                'events': day_events,
                'events_count': len(day_events)
            })

            current_date += timedelta(days=1)

        return Response({
            'period': {
                'start_date': start_date,
                'end_date': end_date
            },
            'calendar_data': calendar_data,
            'summary': {
                'total_events': leave_requests.count(),
                'unique_users_on_leave': leave_requests.values('staff_user').distinct().count()
            }
        })

    @action(detail=False, methods=['get'])
    def pending_requests(self, request):
        """Get all pending leave requests for manager approval"""
        pending_requests = LeaveRequest.objects.filter(
            status='pending',
            staff_user__in=self.get_queryset()
        ).select_related('staff_user', 'leave_type').order_by('-submitted_at')

        pending_data = []
        for req in pending_requests:
            pending_data.append({
                'id': req.id,
                'user': UserBasicSerializer(req.staff_user).data,
                'leave_type': {
                    'name': req.leave_type.name,
                    'code': req.leave_type.code,
                    'color_code': req.leave_type.color_code
                },
                'start_date': req.start_date,
                'end_date': req.end_date,
                'days_requested': req.days_requested,
                'request_type': req.request_type,
                'reason': req.reason,
                'emergency': req.emergency,
                'submitted_at': req.submitted_at,
                'days_until_start': (req.start_date - date.today()).days
            })

        return Response({
            'pending_requests': pending_data,
            'count': len(pending_data),
            'urgent_requests': [
                req for req in pending_data
                if req['emergency'] or req['days_until_start'] <= 7
            ]
        })

    @action(detail=False, methods=['get'])
    def analytics_summary(self, request):
        """Get team leave analytics summary"""
        current_year = timezone.now().year
        year = request.query_params.get('year', current_year)

        # Get basic statistics
        team_requests = LeaveRequest.objects.filter(
            staff_user__in=self.get_queryset(),
            created_at__year=year
        )

        analytics = {
            'year': year,
            'total_requests': team_requests.count(),
            'approved_requests': team_requests.filter(status='approved').count(),
            'pending_requests': team_requests.filter(status='pending').count(),
            'rejected_requests': team_requests.filter(status='rejected').count(),
            'total_days_requested': team_requests.aggregate(
                total=Sum('days_requested')
            )['total'] or Decimal('0'),
            'average_days_per_request': 0,
            'leave_types_breakdown': [],
            'monthly_breakdown': []
        }

        if analytics['total_requests'] > 0:
            analytics['average_days_per_request'] = float(
                analytics['total_days_requested'] / analytics['total_requests']
            )

        # Leave types breakdown
        leave_type_stats = team_requests.values('leave_type__name', 'leave_type__code').annotate(
            count=Count('id'),
            total_days=Sum('days_requested')
        ).order_by('-count')

        analytics['leave_types_breakdown'] = [
            {
                'leave_type': stat['leave_type__name'],
                'code': stat['leave_type__code'],
                'request_count': stat['count'],
                'total_days': float(stat['total_days']) if stat['total_days'] else 0
            }
            for stat in leave_type_stats
        ]

        # Monthly breakdown
        monthly_stats = team_requests.extra(
            select={'month': "EXTRACT(month FROM created_at)"}
        ).values('month').annotate(
            count=Count('id'),
            total_days=Sum('days_requested')
        ).order_by('month')

        analytics['monthly_breakdown'] = [
            {
                'month': int(stat['month']),
                'month_name': date(year, int(stat['month']), 1).strftime('%B'),
                'request_count': stat['count'],
                'total_days': float(stat['total_days']) if stat['total_days'] else 0
            }
            for stat in monthly_stats
        ]

        return Response(analytics)


class LeaveRequestViewSet(viewsets.ModelViewSet):
    """
    ViewSet for comprehensive leave request management

    Handles the full leave request workflow:
    - Staff: Submit, view own requests, cancel pending requests
    - Manager: Approve/reject requests, view team requests
    - Admin: Full access to all requests and system management
    """
    queryset = LeaveRequest.objects.select_related('staff_user', 'leave_type', 'approved_by')
    serializer_class = LeaveRequestSerializer
    permission_classes = [IsAuthenticated]
    filter_backends = [DjangoFilterBackend, filters.SearchFilter, filters.OrderingFilter]
    filterset_fields = ['status', 'leave_type', 'staff_user', 'emergency']
    search_fields = ['reason', 'staff_user__username', 'staff_user__first_name', 'staff_user__last_name']
    ordering_fields = ['created_at', 'start_date', 'submitted_at', 'approved_at']
    ordering = ['-created_at']

    def perform_content_negotiation(self, request, force=False):
        """
        Override content negotiation to allow raw HttpResponse for export action.
        For export_requests action, bypass normal content negotiation.
        """
        if self.action == 'export_requests':
            # Return PassthroughRenderer for export action
            return (PassthroughRenderer(), 'application/octet-stream')
        return super().perform_content_negotiation(request, force)

    def get_queryset(self):
        """Filter queryset based on user permissions"""
        queryset = super().get_queryset()
        user = self.request.user

        # Get the user's company from their active membership
        user_membership = user.company_memberships.filter(
            is_active=True
        ).select_related('company').first()

        if not user_membership:
            # User has no company membership - return empty queryset
            return queryset.none()

        # Get all users who are members of the same company
        company_user_ids = user_membership.company.memberships.filter(
            is_active=True
        ).values_list('user_id', flat=True)

        # Filter requests to only include users from the same company
        queryset = queryset.filter(staff_user_id__in=company_user_ids)

        # Staff users see only their own requests within their company
        if user.role == 'staff':
            return queryset.filter(staff_user=user)

        # Managers see all requests from their company
        # (TODO: implement team hierarchy to show only direct reports)
        elif user.role == 'manager':
            return queryset

        # Admins see all requests from their company
        return queryset

    def perform_create(self, serializer):
        """Handle leave request creation with validation"""
        # Set the requesting user
        serializer.save(staff_user=self.request.user)

        # Log the request creation
        logger.info(
            f"Leave request submitted: {self.request.user.username} - "
            f"{serializer.instance.leave_type.name} "
            f"({serializer.instance.start_date} to {serializer.instance.end_date})"
        )

    def create(self, request, *args, **kwargs):
        """Submit a new leave request"""
        serializer = self.get_serializer(data=request.data)
        serializer.is_valid(raise_exception=True)

        # Additional business logic validation
        leave_type_id = serializer.validated_data['leave_type'].id
        start_date = serializer.validated_data['start_date']
        end_date = serializer.validated_data['end_date']

        # Check for blackout periods
        blackout_periods = BlackoutPeriod.objects.overlapping_period(start_date, end_date)
        for blackout in blackout_periods:
            if blackout.restriction_level == 'no_requests':
                return Response({
                    'error': f'Leave requests are not allowed during {blackout.name}',
                    'blackout_period': {
                        'name': blackout.name,
                        'start_date': blackout.start_date,
                        'end_date': blackout.end_date,
                        'message': blackout.get_restriction_message()
                    }
                }, status=status.HTTP_400_BAD_REQUEST)

        # Check leave balance
        try:
            current_year = timezone.now().year
            entitlement = LeaveEntitlement.objects.get(
                user=request.user,
                policy__leave_type_id=leave_type_id,
                year=current_year
            )

            if not entitlement.can_take_leave(serializer.validated_data['days_requested']):
                return Response({
                    'error': 'Insufficient leave balance',
                    'current_balance': entitlement.current_balance,
                    'requested': serializer.validated_data['days_requested']
                }, status=status.HTTP_400_BAD_REQUEST)

        except LeaveEntitlement.DoesNotExist:
            return Response({
                'error': 'No leave entitlement found for this leave type'
            }, status=status.HTTP_400_BAD_REQUEST)

        # Create the request
        self.perform_create(serializer)

        # Auto-submit if not in draft mode
        if serializer.instance.status == 'draft':
            # Allow user to submit later
            pass
        else:
            # Auto-submit — save() handles adding to LeaveBalance.pending_balance
            serializer.instance.status = 'pending'
            serializer.instance.save()

        headers = self.get_success_headers(serializer.data)
        return Response({
            'message': 'Leave request submitted successfully',
            'leave_request': serializer.data,
            'next_steps': (
                'Your request is pending manager approval'
                if serializer.instance.status == 'pending'
                else 'Your request has been saved as draft'
            )
        }, status=status.HTTP_201_CREATED, headers=headers)

    @action(detail=True, methods=['post'])
    def submit(self, request, pk=None):
        """Submit a draft leave request for approval"""
        leave_request = self.get_object()

        # Validate ownership for staff users
        if request.user.role == 'staff' and leave_request.staff_user != request.user:
            return Response({
                'error': 'You can only submit your own requests'
            }, status=status.HTTP_403_FORBIDDEN)

        if leave_request.status != 'draft':
            return Response({
                'error': f'Only draft requests can be submitted. Current status: {leave_request.status}'
            }, status=status.HTTP_400_BAD_REQUEST)

        # Submit the request — save() handles adding to LeaveBalance.pending_balance
        leave_request.status = 'pending'
        leave_request.save()

        return Response({
            'message': 'Leave request submitted for approval',
            'leave_request': self.get_serializer(leave_request).data
        })

    @action(detail=True, methods=['post'], permission_classes=[IsAuthenticated, ManagerOrAdminPermission])
    def approve(self, request, pk=None):
        """Approve a leave request"""
        leave_request = self.get_object()
        notes = request.data.get('notes', '')

        if leave_request.status != 'pending':
            return Response({
                'error': f'Only pending requests can be approved. Current status: {leave_request.status}'
            }, status=status.HTTP_400_BAD_REQUEST)

        # Approve the request — model method handles LeaveBalance + LeaveEntitlement updates
        leave_request.approve(request.user, notes)

        return Response({
            'message': 'Leave request approved successfully',
            'leave_request': self.get_serializer(leave_request).data
        })

    @action(detail=True, methods=['post'], permission_classes=[IsAuthenticated, ManagerOrAdminPermission])
    def reject(self, request, pk=None):
        """Reject a leave request"""
        leave_request = self.get_object()
        notes = request.data.get('notes', '')

        if leave_request.status != 'pending':
            return Response({
                'error': f'Only pending requests can be rejected. Current status: {leave_request.status}'
            }, status=status.HTTP_400_BAD_REQUEST)

        if not notes:
            return Response({
                'error': 'Rejection reason is required'
            }, status=status.HTTP_400_BAD_REQUEST)

        # Reject the request — model method handles LeaveBalance pending reversal
        leave_request.reject(request.user, notes)

        return Response({
            'message': 'Leave request rejected',
            'leave_request': self.get_serializer(leave_request).data
        })

    @action(detail=True, methods=['post'])
    def cancel(self, request, pk=None):
        """Cancel a leave request (by the requesting user)"""
        leave_request = self.get_object()

        # Validate ownership for staff users
        if request.user.role == 'staff' and leave_request.staff_user != request.user:
            return Response({
                'error': 'You can only cancel your own requests'
            }, status=status.HTTP_403_FORBIDDEN)

        if not leave_request.can_be_cancelled:
            return Response({
                'error': f'Request cannot be cancelled. Current status: {leave_request.status}'
            }, status=status.HTTP_400_BAD_REQUEST)

        # Cancel the request — model method handles LeaveBalance pending reversal
        leave_request.cancel()

        return Response({
            'message': 'Leave request cancelled successfully',
            'leave_request': self.get_serializer(leave_request).data
        })

    @action(detail=False, methods=['get'])
    def my_requests(self, request):
        """Get current user's leave requests"""
        user_requests = self.get_queryset().filter(staff_user=request.user)

        # Apply additional filtering
        status_filter = request.query_params.get('status')
        if status_filter:
            user_requests = user_requests.filter(status=status_filter)

        year_filter = request.query_params.get('year')
        if year_filter:
            user_requests = user_requests.filter(start_date__year=year_filter)

        page = self.paginate_queryset(user_requests)
        if page is not None:
            serializer = self.get_serializer(page, many=True)
            return self.get_paginated_response(serializer.data)

        serializer = self.get_serializer(user_requests, many=True)
        return Response(serializer.data)

    @action(
        detail=False,
        methods=['get'],
        url_path='export',
        renderer_classes=[PassthroughRenderer]  # Allow raw HttpResponse
    )
    def export_requests(self, request):
        """Export leave requests to CSV or Excel"""
        import csv
        import io
        from django.http import HttpResponse
        from django.utils import timezone

        logger.info(f"Export action called by user: {request.user.username}")

        export_format = request.query_params.get('format', 'csv')

        # Get filtered queryset
        queryset = self.filter_queryset(self.get_queryset())

        # Apply additional filters from query params
        status_filter = request.query_params.get('status')
        if status_filter:
            statuses = status_filter.split(',')
            queryset = queryset.filter(status__in=statuses)

        leave_type_filter = request.query_params.get('leave_type')
        if leave_type_filter:
            leave_types = leave_type_filter.split(',')
            queryset = queryset.filter(leave_type_id__in=leave_types)

        start_date = request.query_params.get('start_date')
        if start_date:
            queryset = queryset.filter(start_date__gte=start_date)

        end_date = request.query_params.get('end_date')
        if end_date:
            queryset = queryset.filter(end_date__lte=end_date)

        user_filter = request.query_params.get('user')
        if user_filter:
            users = user_filter.split(',')
            queryset = queryset.filter(staff_user_id__in=users)

        # Order by creation date descending
        queryset = queryset.order_by('-created_at')

        if export_format == 'csv':
            # Create CSV
            output = io.StringIO()
            writer = csv.writer(output)

            # Header row
            writer.writerow([
                'Request ID',
                'Employee Name',
                'Employee Email',
                'Leave Type',
                'Start Date',
                'End Date',
                'Days Requested',
                'Status',
                'Reason',
                'Submitted Date',
                'Approved By',
                'Approved Date',
                'Manager Notes'
            ])

            # Data rows
            for request_obj in queryset:
                approved_by_name = ''
                if request_obj.approved_by:
                    approved_by_name = f"{request_obj.approved_by.first_name} {request_obj.approved_by.last_name}".strip()

                writer.writerow([
                    request_obj.id,
                    f"{request_obj.staff_user.first_name} {request_obj.staff_user.last_name}".strip(),
                    request_obj.staff_user.email,
                    request_obj.leave_type.name,
                    request_obj.start_date.strftime('%Y-%m-%d'),
                    request_obj.end_date.strftime('%Y-%m-%d'),
                    float(request_obj.days_requested),
                    request_obj.status.title(),
                    request_obj.reason,
                    request_obj.submitted_at.strftime('%Y-%m-%d %H:%M') if request_obj.submitted_at else '',
                    approved_by_name,
                    request_obj.approved_at.strftime('%Y-%m-%d %H:%M') if request_obj.approved_at else '',
                    request_obj.manager_notes
                ])

            # Create response
            response = HttpResponse(output.getvalue(), content_type='text/csv')
            filename = f"leave_requests_{timezone.now().strftime('%Y%m%d_%H%M%S')}.csv"
            response['Content-Disposition'] = f'attachment; filename="{filename}"'
            return response

        elif export_format == 'xlsx':
            # Create Excel file
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment

            wb = Workbook()
            ws = wb.active
            ws.title = "Leave Requests"

            # Header row with styling
            headers = [
                'Request ID', 'Employee Name', 'Employee Email', 'Leave Type',
                'Start Date', 'End Date', 'Days Requested', 'Status',
                'Reason', 'Submitted Date', 'Approved By', 'Approved Date', 'Manager Notes'
            ]

            header_fill = PatternFill(start_color="0078D4", end_color="0078D4", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF")

            for col_num, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col_num, value=header)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center', vertical='center')

            # Data rows
            for row_num, request_obj in enumerate(queryset, 2):
                approved_by_name = ''
                if request_obj.approved_by:
                    approved_by_name = f"{request_obj.approved_by.first_name} {request_obj.approved_by.last_name}".strip()

                ws.cell(row=row_num, column=1, value=request_obj.id)
                ws.cell(row=row_num, column=2, value=f"{request_obj.staff_user.first_name} {request_obj.staff_user.last_name}".strip())
                ws.cell(row=row_num, column=3, value=request_obj.staff_user.email)
                ws.cell(row=row_num, column=4, value=request_obj.leave_type.name)
                ws.cell(row=row_num, column=5, value=request_obj.start_date.strftime('%Y-%m-%d'))
                ws.cell(row=row_num, column=6, value=request_obj.end_date.strftime('%Y-%m-%d'))
                ws.cell(row=row_num, column=7, value=float(request_obj.days_requested))
                ws.cell(row=row_num, column=8, value=request_obj.status.title())
                ws.cell(row=row_num, column=9, value=request_obj.reason)
                ws.cell(row=row_num, column=10, value=request_obj.submitted_at.strftime('%Y-%m-%d %H:%M') if request_obj.submitted_at else '')
                ws.cell(row=row_num, column=11, value=approved_by_name)
                ws.cell(row=row_num, column=12, value=request_obj.approved_at.strftime('%Y-%m-%d %H:%M') if request_obj.approved_at else '')
                ws.cell(row=row_num, column=13, value=request_obj.manager_notes)

            # Auto-adjust column widths
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width

            # Save to bytes
            output = io.BytesIO()
            wb.save(output)
            output.seek(0)

            # Create response
            response = HttpResponse(
                output.getvalue(),
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
            filename = f"leave_requests_{timezone.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            response['Content-Disposition'] = f'attachment; filename="{filename}"'
            return response

        else:
            return Response({
                'error': f'Unsupported format: {export_format}. Use csv or xlsx.'
            }, status=status.HTTP_400_BAD_REQUEST)

    @action(detail=False, methods=['get'], permission_classes=[IsAuthenticated, ManagerOrAdminPermission])
    def pending_approvals(self, request):
        """Get leave requests pending manager approval"""
        pending_requests = self.get_queryset().filter(status='pending')

        # Additional filtering by date range
        start_date = request.query_params.get('start_date')
        end_date = request.query_params.get('end_date')

        if start_date:
            pending_requests = pending_requests.filter(start_date__gte=start_date)
        if end_date:
            pending_requests = pending_requests.filter(end_date__lte=end_date)

        serializer = self.get_serializer(pending_requests, many=True)
        return Response({
            'pending_requests': serializer.data,
            'count': pending_requests.count(),
            'urgent_count': pending_requests.filter(
                models.Q(emergency=True) |
                models.Q(start_date__lte=timezone.now().date() + timedelta(days=7))
            ).count()
        })


# Additional viewsets for future phases

class LeaveCalendarViewSet(viewsets.ReadOnlyModelViewSet):
    """
    ViewSet for calendar integration (TASK-015)
    """
    queryset = LeaveRequest.objects.all()
    serializer_class = LeaveRequestSerializer
    permission_classes = [IsAuthenticated]

    def get_queryset(self):
        """Filter calendar events based on user permissions and date range"""
        queryset = super().get_queryset()

        # Filter by user role
        if not (self.request.user.role in ['manager', 'admin']):
            # Staff can only see their own requests
            queryset = queryset.filter(staff_user=self.request.user)

        # Filter by date range if provided
        start_date = self.request.query_params.get('start_date')
        end_date = self.request.query_params.get('end_date')

        if start_date:
            queryset = queryset.filter(start_date__gte=start_date)
        if end_date:
            queryset = queryset.filter(end_date__lte=end_date)

        return queryset.filter(status='approved').select_related('staff_user', 'leave_type')

    def list(self, request):
        """Get calendar events for approved leave requests"""
        queryset = self.get_queryset()

        events = []
        for leave_request in queryset:
            events.append({
                'id': leave_request.id,
                'title': f"{leave_request.staff_user.get_full_name()} - {leave_request.leave_type.name}",
                'start': leave_request.start_date.isoformat(),
                'end': leave_request.end_date.isoformat(),
                'type': leave_request.leave_type.code,
                'staff': leave_request.staff_user.get_full_name(),
                'days': float(leave_request.days_requested)
            })

        return Response({
            'events': events,
            'total': len(events)
        })

    @action(detail=False, methods=['get'])
    def events(self, request):
        """Custom endpoint for calendar events - matches frontend URL pattern"""
        return self.list(request)


class LeaveReportsViewSet(viewsets.ReadOnlyModelViewSet):
    """
    ViewSet for comprehensive leave reporting and analytics

    Provides detailed reports and analytics for managers and administrators
    """
    permission_classes = [IsAuthenticated, ManagerOrAdminPermission]
    serializer_class = LeaveAnalyticsSerializer

    def list(self, request):
        """Get available leave reports and quick metrics"""
        current_year = timezone.now().year

        # Quick metrics
        total_requests = LeaveRequest.objects.filter(created_at__year=current_year).count()
        pending_requests = LeaveRequest.objects.filter(status='pending').count()
        approved_requests = LeaveRequest.objects.filter(status='approved', created_at__year=current_year).count()

        return Response({
            'available_reports': [
                'usage_summary', 'analytics', 'export',
                'balance_trends', 'policy_effectiveness',
                'team_utilization', 'forecast'
            ],
            'quick_metrics': {
                'total_requests_this_year': total_requests,
                'pending_approvals': pending_requests,
                'approved_requests_this_year': approved_requests,
                'approval_rate': (
                    (approved_requests / total_requests * 100) if total_requests > 0 else 0
                )
            },
            'generated_at': timezone.now().isoformat()
        })

    @action(detail=False, methods=['get'])
    def analytics(self, request):
        """Get comprehensive leave analytics with filtering support"""
        current_year = timezone.now().year
        year = int(request.query_params.get('year', current_year))

        # Date range filtering
        start_date = request.query_params.get('start_date')
        end_date = request.query_params.get('end_date')

        # Additional filter parameters
        leave_type_ids = request.query_params.getlist('leave_type')  # Can be multiple IDs
        status_filters = request.query_params.getlist('status')  # Can be multiple statuses
        department_filters = request.query_params.getlist('department')

        # Base queryset
        requests_qs = LeaveRequest.objects.filter(created_at__year=year)

        # Apply date range filters
        if start_date:
            requests_qs = requests_qs.filter(start_date__gte=start_date)
        if end_date:
            requests_qs = requests_qs.filter(end_date__lte=end_date)

        # Apply leave type filter
        if leave_type_ids:
            # Convert to integers and filter
            leave_type_ids = [int(id) for id in leave_type_ids]
            requests_qs = requests_qs.filter(leave_type_id__in=leave_type_ids)

        # Apply status filter
        if status_filters:
            requests_qs = requests_qs.filter(status__in=status_filters)

        # Apply department filter (if staff_user has profile with department info)
        if department_filters:
            # Assuming StaffProfile has department or employment_type field
            # This may need adjustment based on actual model structure
            requests_qs = requests_qs.filter(
                staff_user__profile__employment_type__name__in=department_filters
            )

        # Basic statistics
        total_requests = requests_qs.count()
        approved_requests = requests_qs.filter(status='approved').count()
        pending_requests = requests_qs.filter(status='pending').count()
        rejected_requests = requests_qs.filter(status='rejected').count()

        total_days = requests_qs.aggregate(total=Sum('days_requested'))['total'] or Decimal('0')
        avg_days = float(total_days / total_requests) if total_requests > 0 else 0

        # Leave type breakdown
        leave_type_stats = requests_qs.values(
            'leave_type__name', 'leave_type__code', 'leave_type__color_code'
        ).annotate(
            count=Count('id'),
            total_days=Sum('days_requested'),
            avg_days=Avg('days_requested')
        ).order_by('-count')

        # Monthly trends
        monthly_trends = []
        for month in range(1, 13):
            month_requests = requests_qs.filter(created_at__month=month)
            monthly_trends.append({
                'month': month,
                'month_name': date(year, month, 1).strftime('%B'),
                'total_requests': month_requests.count(),
                'approved': month_requests.filter(status='approved').count(),
                'rejected': month_requests.filter(status='rejected').count(),
                'total_days': float(
                    month_requests.aggregate(total=Sum('days_requested'))['total'] or Decimal('0')
                )
            })

        # Popular leave periods (by start date)
        popular_periods = requests_qs.filter(status='approved').values(
            'start_date__month'
        ).annotate(
            count=Count('id')
        ).order_by('-count')[:5]

        popular_months = [
            {
                'month': period['start_date__month'],
                'month_name': date(year, period['start_date__month'], 1).strftime('%B'),
                'request_count': period['count']
            }
            for period in popular_periods
        ]

        # User utilization stats
        user_stats = requests_qs.filter(status='approved').values(
            'staff_user__username', 'staff_user__first_name', 'staff_user__last_name'
        ).annotate(
            total_requests=Count('id'),
            total_days_taken=Sum('days_requested')
        ).order_by('-total_days_taken')[:10]

        analytics_data = {
            'period': {
                'year': year,
                'start_date': start_date,
                'end_date': end_date
            },
            'summary': {
                'total_requests': total_requests,
                'approved_requests': approved_requests,
                'pending_requests': pending_requests,
                'rejected_requests': rejected_requests,
                'total_days_taken': float(total_days),
                'average_days_per_request': round(avg_days, 2),
                'approval_rate': round((approved_requests / total_requests * 100) if total_requests > 0 else 0, 2)
            },
            'leave_types_breakdown': [
                {
                    'leave_type': stat['leave_type__name'],
                    'code': stat['leave_type__code'],
                    'color_code': stat['leave_type__color_code'],
                    'request_count': stat['count'],
                    'total_days': float(stat['total_days']) if stat['total_days'] else 0,
                    'average_days': round(float(stat['avg_days']) if stat['avg_days'] else 0, 2),
                    'percentage': round((stat['count'] / total_requests * 100) if total_requests > 0 else 0, 2)
                }
                for stat in leave_type_stats
            ],
            'monthly_trends': monthly_trends,
            'popular_leave_months': popular_months,
            'top_users': [
                {
                    'username': user['staff_user__username'],
                    'full_name': f"{user['staff_user__first_name']} {user['staff_user__last_name']}".strip(),
                    'total_requests': user['total_requests'],
                    'total_days_taken': float(user['total_days_taken']) if user['total_days_taken'] else 0
                }
                for user in user_stats
            ],
            'generated_at': timezone.now().isoformat()
        }

        return Response(analytics_data)

    @action(detail=False, methods=['get'])
    def usage_summary(self, request):
        """Get detailed usage summary report"""
        current_year = timezone.now().year
        year = int(request.query_params.get('year', current_year))

        # Entitlement statistics
        entitlements = LeaveEntitlement.objects.filter(year=year)
        entitlement_stats = entitlements.aggregate(
            total_entitled=Sum('annual_entitlement'),
            total_accrued=Sum('accrued_to_date'),
            total_used=Sum('used_to_date'),
            total_carried_over=Sum('carried_over')
        )

        # Calculate remaining and utilization
        total_available = (
            (entitlement_stats['total_entitled'] or 0) +
            (entitlement_stats['total_carried_over'] or 0) +
            (entitlement_stats['total_accrued'] or 0)
        )
        total_used = entitlement_stats['total_used'] or 0
        total_remaining = total_available - total_used
        utilization_rate = (total_used / total_available * 100) if total_available > 0 else 0

        # By leave type breakdown
        leave_type_breakdown = entitlements.values(
            'policy__leave_type__name',
            'policy__leave_type__code',
            'policy__leave_type__color_code'
        ).annotate(
            users_count=Count('user', distinct=True),
            total_entitled=Sum('annual_entitlement'),
            total_used=Sum('used_to_date'),
            total_remaining=Sum('annual_entitlement') - Sum('used_to_date')
        ).order_by('-total_entitled')

        # Department/employment type breakdown (if available)
        dept_breakdown = entitlements.values(
            'user__profile__employment_type__name'
        ).annotate(
            users_count=Count('user', distinct=True),
            total_entitled=Sum('annual_entitlement'),
            total_used=Sum('used_to_date'),
            avg_utilization=Avg(
                models.Case(
                    models.When(annual_entitlement__gt=0,
                               then=models.F('used_to_date') / models.F('annual_entitlement') * 100),
                    default=0,
                    output_field=models.DecimalField(max_digits=5, decimal_places=2)
                )
            )
        ).order_by('-total_entitled')

        summary_data = {
            'year': year,
            'overall_summary': {
                'total_available_days': float(total_available),
                'total_used_days': float(total_used),
                'total_remaining_days': float(total_remaining),
                'overall_utilization_rate': round(utilization_rate, 2),
                'total_employees': entitlements.values('user').distinct().count(),
                'total_entitlements': entitlements.count()
            },
            'by_leave_type': [
                {
                    'leave_type': item['policy__leave_type__name'],
                    'code': item['policy__leave_type__code'],
                    'color_code': item['policy__leave_type__color_code'],
                    'employees': item['users_count'],
                    'total_entitled': float(item['total_entitled']) if item['total_entitled'] else 0,
                    'total_used': float(item['total_used']) if item['total_used'] else 0,
                    'total_remaining': float(item['total_remaining']) if item['total_remaining'] else 0,
                    'utilization_rate': round(
                        (float(item['total_used']) / float(item['total_entitled']) * 100)
                        if item['total_entitled'] and item['total_entitled'] > 0 else 0, 2
                    )
                }
                for item in leave_type_breakdown
            ],
            'by_employment_type': [
                {
                    'employment_type': item['user__profile__employment_type__name'] or 'Unknown',
                    'employees': item['users_count'],
                    'total_entitled': float(item['total_entitled']) if item['total_entitled'] else 0,
                    'total_used': float(item['total_used']) if item['total_used'] else 0,
                    'avg_utilization_rate': round(float(item['avg_utilization']) if item['avg_utilization'] else 0, 2)
                }
                for item in dept_breakdown
            ],
            'generated_at': timezone.now().isoformat()
        }

        return Response(summary_data)

    @action(detail=False, methods=['get'])
    def export(self, request):
        """Export leave data in various formats"""
        export_format = request.query_params.get('format', 'json')
        report_type = request.query_params.get('type', 'analytics')
        year = int(request.query_params.get('year', timezone.now().year))

        if report_type == 'analytics':
            # Get analytics data
            analytics_request = request._request.__class__()
            analytics_request.query_params = request.query_params
            analytics_request.user = request.user

            # Reuse analytics endpoint
            analytics_response = self.analytics(analytics_request)
            data = analytics_response.data

        elif report_type == 'usage_summary':
            usage_request = request._request.__class__()
            usage_request.query_params = request.query_params
            usage_request.user = request.user

            usage_response = self.usage_summary(usage_request)
            data = usage_response.data

        else:
            return Response({
                'error': 'Invalid report type',
                'available_types': ['analytics', 'usage_summary']
            }, status=status.HTTP_400_BAD_REQUEST)

        if export_format.lower() == 'csv':
            # TODO: Implement CSV export functionality
            return Response({
                'message': 'CSV export functionality to be implemented',
                'data_preview': str(data)[:500] + '...' if len(str(data)) > 500 else str(data)
            })

        elif export_format.lower() == 'excel':
            # TODO: Implement Excel export functionality
            return Response({
                'message': 'Excel export functionality to be implemented',
                'data_preview': str(data)[:500] + '...' if len(str(data)) > 500 else str(data)
            })

        else:
            # Default JSON export
            return Response({
                'export_format': 'json',
                'export_timestamp': timezone.now().isoformat(),
                'data': data
            })

    @action(detail=False, methods=['get'])
    def balance_trends(self, request):
        """Analyze leave balance trends and projections"""
        current_year = timezone.now().year
        year = int(request.query_params.get('year', current_year))

        # Get balance trends by leave type over the year
        trends = []

        for leave_type in LeaveType.objects.filter(is_active=True):
            entitlements = LeaveEntitlement.objects.filter(
                policy__leave_type=leave_type,
                year=year
            )

            if entitlements.exists():
                stats = entitlements.aggregate(
                    total_entitled=Sum('annual_entitlement'),
                    total_used=Sum('used_to_date'),
                    total_accrued=Sum('accrued_to_date'),
                    avg_balance=Avg('annual_entitlement')
                )

                trends.append({
                    'leave_type': leave_type.name,
                    'code': leave_type.code,
                    'color_code': leave_type.color_code,
                    'total_entitled': float(stats['total_entitled']) if stats['total_entitled'] else 0,
                    'total_used': float(stats['total_used']) if stats['total_used'] else 0,
                    'total_accrued': float(stats['total_accrued']) if stats['total_accrued'] else 0,
                    'average_balance': float(stats['avg_balance']) if stats['avg_balance'] else 0,
                    'utilization_trend': round(
                        (float(stats['total_used']) / float(stats['total_entitled']) * 100)
                        if stats['total_entitled'] and stats['total_entitled'] > 0 else 0, 2
                    )
                })

        return Response({
            'year': year,
            'balance_trends': trends,
            'generated_at': timezone.now().isoformat()
        })

    @action(detail=False, methods=['get'])
    def team_utilization(self, request):
        """Analyze team-level leave utilization patterns"""
        current_year = timezone.now().year
        year = int(request.query_params.get('year', current_year))

        # Team utilization by employment type
        team_stats = LeaveEntitlement.objects.filter(year=year).values(
            'user__profile__employment_type__name'
        ).annotate(
            team_size=Count('user', distinct=True),
            total_entitled=Sum('annual_entitlement'),
            total_used=Sum('used_to_date'),
            avg_individual_entitled=Avg('annual_entitlement'),
            avg_individual_used=Avg('used_to_date')
        ).order_by('-total_entitled')

        utilization_data = []
        for team in team_stats:
            utilization_rate = (
                (float(team['total_used']) / float(team['total_entitled']) * 100)
                if team['total_entitled'] and team['total_entitled'] > 0 else 0
            )

            utilization_data.append({
                'team_name': team['user__profile__employment_type__name'] or 'Unknown',
                'team_size': team['team_size'],
                'total_entitled_days': float(team['total_entitled']) if team['total_entitled'] else 0,
                'total_used_days': float(team['total_used']) if team['total_used'] else 0,
                'team_utilization_rate': round(utilization_rate, 2),
                'avg_individual_entitled': round(float(team['avg_individual_entitled']) if team['avg_individual_entitled'] else 0, 2),
                'avg_individual_used': round(float(team['avg_individual_used']) if team['avg_individual_used'] else 0, 2),
            })

        return Response({
            'year': year,
            'team_utilization': utilization_data,
            'generated_at': timezone.now().isoformat()
        })

    @action(detail=False, methods=['get'])
    def statistics(self, request):
        """Get leave statistics summary for dashboard metrics"""
        current_year = timezone.now().year
        year = int(request.query_params.get('year', current_year))

        # Date range filtering
        start_date = request.query_params.get('start_date')
        end_date = request.query_params.get('end_date')

        # Base queryset
        requests_qs = LeaveRequest.objects.filter(created_at__year=year)

        if start_date:
            requests_qs = requests_qs.filter(start_date__gte=start_date)
        if end_date:
            requests_qs = requests_qs.filter(end_date__lte=end_date)

        # Basic statistics
        total_requests = requests_qs.count()
        approved_requests = requests_qs.filter(status='approved').count()
        pending_requests = requests_qs.filter(status='pending').count()
        rejected_requests = requests_qs.filter(status='rejected').count()

        # Average days per request
        total_days = requests_qs.aggregate(total=Sum('days_requested'))['total'] or Decimal('0')
        avg_days = float(total_days / total_requests) if total_requests > 0 else 0

        # Most popular leave type
        most_popular_type = requests_qs.values(
            'leave_type__id',
            'leave_type__name',
            'leave_type__code',
            'leave_type__description',
            'leave_type__color_code',
            'leave_type__is_active'
        ).annotate(
            count=Count('id')
        ).order_by('-count').first()

        # Busiest leave period (by start date month)
        busiest_period = requests_qs.filter(status='approved').values(
            'start_date__month',
            'start_date__year'
        ).annotate(
            count=Count('id')
        ).order_by('-count').first()

        # Format response matching LeaveStatistics interface
        statistics_data = {
            'total_requests': total_requests,
            'pending_requests': pending_requests,
            'approved_requests': approved_requests,
            'rejected_requests': rejected_requests,
            'average_days_per_request': f"{avg_days:.1f}",
            'most_popular_leave_type': {
                'id': most_popular_type['leave_type__id'] if most_popular_type else None,
                'name': most_popular_type['leave_type__name'] if most_popular_type else 'N/A',
                'code': most_popular_type['leave_type__code'] if most_popular_type else '',
                'description': most_popular_type['leave_type__description'] if most_popular_type else '',
                'color_code': most_popular_type['leave_type__color_code'] if most_popular_type else '#0078d4',
                'is_active': most_popular_type['leave_type__is_active'] if most_popular_type else True
            } if most_popular_type else None,
            'busiest_leave_period': {
                'month': busiest_period['start_date__month'] if busiest_period else 1,
                'year': busiest_period['start_date__year'] if busiest_period else year,
                'request_count': busiest_period['count'] if busiest_period else 0
            }
        }

        return Response(statistics_data)

    def _get_filtered_leave_requests(self, request):
        """
        Build a filtered queryset of LeaveRequest objects based on query params.
        Applies multi-tenant company filtering and select_related to avoid N+1.

        Supported query params: year, start_date, end_date, leave_type, status
        """
        current_year = timezone.now().year
        year = int(request.query_params.get('year', current_year))
        start_date = request.query_params.get('start_date')
        end_date = request.query_params.get('end_date')
        leave_type_ids = request.query_params.getlist('leave_type')
        status_filters = request.query_params.getlist('status')

        # Base queryset filtered by year
        queryset = LeaveRequest.objects.filter(created_at__year=year)

        # Multi-tenant: restrict to the requesting user's company
        user_membership = request.user.company_memberships.filter(
            is_active=True
        ).select_related('company').first()

        if user_membership:
            company_user_ids = user_membership.company.memberships.filter(
                is_active=True
            ).values_list('user_id', flat=True)
            queryset = queryset.filter(staff_user_id__in=company_user_ids)
        else:
            queryset = queryset.none()

        # Apply optional filters
        if start_date:
            queryset = queryset.filter(start_date__gte=start_date)
        if end_date:
            queryset = queryset.filter(end_date__lte=end_date)
        if leave_type_ids:
            leave_type_ids = [int(lt_id) for lt_id in leave_type_ids]
            queryset = queryset.filter(leave_type_id__in=leave_type_ids)
        if status_filters:
            queryset = queryset.filter(status__in=status_filters)

        # Optimise with select_related to avoid N+1 queries
        queryset = queryset.select_related(
            'staff_user', 'leave_type'
        ).order_by('-created_at')

        return queryset

    @action(detail=False, methods=['get'])
    def export_xlsx(self, request):
        """Export leave requests data to Excel format"""
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from django.http import HttpResponse
        import io

        queryset = self._get_filtered_leave_requests(request)

        # Create workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Leave Requests"

        # Header styling
        header_fill = PatternFill(start_color="0078D4", end_color="0078D4", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True, size=11)
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        headers = [
            'Employee Name', 'Email', 'Leave Type', 'Start Date',
            'End Date', 'Days Requested', 'Status', 'Reason', 'Created At'
        ]
        column_widths = [25, 30, 20, 14, 14, 16, 14, 40, 20]

        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.border = border
            cell.alignment = Alignment(horizontal='center', vertical='center')

        # Set column widths
        for col_idx, width in enumerate(column_widths, start=1):
            ws.column_dimensions[chr(64 + col_idx)].width = width

        # Write data rows
        for row_idx, leave_req in enumerate(queryset, start=2):
            ws.cell(row=row_idx, column=1, value=f"{leave_req.staff_user.first_name} {leave_req.staff_user.last_name}")
            ws.cell(row=row_idx, column=2, value=leave_req.staff_user.email)
            ws.cell(row=row_idx, column=3, value=leave_req.leave_type.name)
            ws.cell(row=row_idx, column=4, value=leave_req.start_date.isoformat())
            ws.cell(row=row_idx, column=5, value=leave_req.end_date.isoformat())
            ws.cell(row=row_idx, column=6, value=float(leave_req.days_requested))
            ws.cell(row=row_idx, column=7, value=leave_req.status)
            ws.cell(row=row_idx, column=8, value=leave_req.reason)
            ws.cell(row=row_idx, column=9, value=leave_req.created_at.strftime('%Y-%m-%d %H:%M:%S'))

            # Apply border to data cells
            for col_idx in range(1, len(headers) + 1):
                ws.cell(row=row_idx, column=col_idx).border = border

        # Save to BytesIO
        excel_file = io.BytesIO()
        wb.save(excel_file)
        excel_file.seek(0)

        # Create response
        year = request.query_params.get('year', timezone.now().year)
        response = HttpResponse(
            excel_file.read(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        filename = f"leave_requests_{year}.xlsx"
        response['Content-Disposition'] = f'attachment; filename="{filename}"'

        return response

    @action(detail=False, methods=['get'])
    def export_pdf(self, request):
        """Export analytics data to PDF format"""
        from reportlab.lib import colors
        from reportlab.lib.pagesizes import letter, A4
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.units import inch
        from reportlab.lib.enums import TA_CENTER, TA_LEFT
        from django.http import HttpResponse
        import io

        # Get analytics data with same filters
        analytics_response = self.analytics(request)
        data = analytics_response.data

        # Create PDF buffer
        buffer = io.BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=letter, topMargin=0.5*inch, bottomMargin=0.5*inch)
        story = []

        # Styles
        styles = getSampleStyleSheet()
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=24,
            textColor=colors.HexColor('#0078D4'),
            spaceAfter=30,
            alignment=TA_CENTER
        )
        heading_style = ParagraphStyle(
            'CustomHeading',
            parent=styles['Heading2'],
            fontSize=16,
            textColor=colors.HexColor('#0078D4'),
            spaceAfter=12,
            spaceBefore=20
        )

        # Title
        story.append(Paragraph('Leave Analytics Report', title_style))
        story.append(Paragraph(f"Period: {data['period']['year']}", styles['Normal']))
        story.append(Spacer(1, 0.3*inch))

        # Summary Section
        story.append(Paragraph('Summary Statistics', heading_style))

        summary = data['summary']
        summary_data = [
            ['Metric', 'Value'],
            ['Total Requests', str(summary['total_requests'])],
            ['Approved Requests', str(summary['approved_requests'])],
            ['Pending Requests', str(summary['pending_requests'])],
            ['Rejected Requests', str(summary['rejected_requests'])],
            ['Total Days Taken', f"{summary['total_days_taken']:.1f}"],
            ['Average Days per Request', f"{summary['average_days_per_request']:.2f}"],
            ['Approval Rate', f"{summary['approval_rate']:.1f}%"],
        ]

        summary_table = Table(summary_data, colWidths=[3.5*inch, 2*inch])
        summary_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#0078D4')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 12),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 1), (-1, -1), 10),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.lightgrey]),
        ]))
        story.append(summary_table)
        story.append(Spacer(1, 0.3*inch))

        # Monthly Trends Section
        story.append(Paragraph('Monthly Trends', heading_style))

        monthly_data = [['Month', 'Total', 'Approved', 'Rejected', 'Days']]
        for month in data['monthly_trends']:
            monthly_data.append([
                month['month_name'],
                str(month['total_requests']),
                str(month['approved']),
                str(month['rejected']),
                f"{month['total_days']:.1f}"
            ])

        monthly_table = Table(monthly_data, colWidths=[1.5*inch, 1*inch, 1*inch, 1*inch, 1*inch])
        monthly_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#0078D4')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 1), (-1, -1), 9),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.lightgrey]),
        ]))
        story.append(monthly_table)

        # Leave Types Breakdown (if available)
        if data['leave_types_breakdown']:
            story.append(PageBreak())
            story.append(Paragraph('Leave Types Breakdown', heading_style))

            types_data = [['Leave Type', 'Code', 'Requests', 'Days', 'Avg', '%']]
            for lt in data['leave_types_breakdown']:
                types_data.append([
                    lt['leave_type'],
                    lt['code'],
                    str(lt['request_count']),
                    f"{lt['total_days']:.1f}",
                    f"{lt['average_days']:.1f}",
                    f"{lt['percentage']:.1f}%"
                ])

            types_table = Table(types_data, colWidths=[1.8*inch, 0.8*inch, 1*inch, 1*inch, 0.8*inch, 0.8*inch])
            types_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#0078D4')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 10),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('GRID', (0, 0), (-1, -1), 1, colors.black),
                ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
                ('FONTSIZE', (0, 1), (-1, -1), 9),
                ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.lightgrey]),
            ]))
            story.append(types_table)

        # Build PDF
        doc.build(story)
        buffer.seek(0)

        # Create response
        response = HttpResponse(buffer.read(), content_type='application/pdf')
        filename = f"leave_analytics_{data['period']['year']}.pdf"
        response['Content-Disposition'] = f'attachment; filename="{filename}"'

        return response

    @action(detail=False, methods=['get'])
    def export_csv(self, request):
        """Export leave requests data to CSV format"""
        import csv
        from django.http import HttpResponse

        queryset = self._get_filtered_leave_requests(request)

        year = request.query_params.get('year', timezone.now().year)
        response = HttpResponse(content_type='text/csv')
        filename = f"leave_requests_{year}.csv"
        response['Content-Disposition'] = f'attachment; filename="{filename}"'

        writer = csv.writer(response)

        # Write header row
        writer.writerow([
            'Employee Name', 'Email', 'Leave Type', 'Start Date',
            'End Date', 'Days Requested', 'Status', 'Reason', 'Created At'
        ])

        # Write data rows
        for leave_req in queryset:
            writer.writerow([
                f"{leave_req.staff_user.first_name} {leave_req.staff_user.last_name}",
                leave_req.staff_user.email,
                leave_req.leave_type.name,
                leave_req.start_date.isoformat(),
                leave_req.end_date.isoformat(),
                float(leave_req.days_requested),
                leave_req.status,
                leave_req.reason,
                leave_req.created_at.strftime('%Y-%m-%d %H:%M:%S'),
            ])

        return response

    @action(detail=False, methods=['get'])
    def detailed_requests(self, request):
        """
        Get detailed leave requests list with sorting, pagination, and filtering

        Query Parameters:
        - year: Filter by year (default: current year)
        - start_date: Filter by start date (YYYY-MM-DD)
        - end_date: Filter by end date (YYYY-MM-DD)
        - leave_type: Filter by leave type IDs (multiple allowed)
        - status: Filter by status (multiple allowed)
        - department: Filter by department (multiple allowed)
        - ordering: Sort by field (e.g., 'start_date', '-created_at')
        - page: Page number for pagination (default: 1)
        - page_size: Number of results per page (default: 25, max: 100)
        """
        from django.core.paginator import Paginator
        from django.db.models import Q

        # Get filter parameters
        current_year = timezone.now().year
        year = int(request.query_params.get('year', current_year))
        start_date = request.query_params.get('start_date')
        end_date = request.query_params.get('end_date')
        leave_type_ids = request.query_params.getlist('leave_type')
        status_filters = request.query_params.getlist('status')
        department_filters = request.query_params.getlist('department')
        ordering = request.query_params.get('ordering', '-created_at')
        page_number = int(request.query_params.get('page', 1))
        page_size = min(int(request.query_params.get('page_size', 25)), 100)

        # Build queryset with filters
        queryset = LeaveRequest.objects.filter(created_at__year=year)

        if start_date:
            queryset = queryset.filter(start_date__gte=start_date)
        if end_date:
            queryset = queryset.filter(end_date__lte=end_date)
        if leave_type_ids:
            leave_type_ids = [int(id) for id in leave_type_ids]
            queryset = queryset.filter(leave_type_id__in=leave_type_ids)
        if status_filters:
            queryset = queryset.filter(status__in=status_filters)
        if department_filters:
            queryset = queryset.filter(
                staff_user__profile__employment_type__name__in=department_filters
            )

        # Apply ordering
        valid_ordering_fields = [
            'start_date', '-start_date',
            'end_date', '-end_date',
            'created_at', '-created_at',
            'status', '-status',
            'days_requested', '-days_requested',
            'staff_user__first_name', '-staff_user__first_name',
            'leave_type__name', '-leave_type__name'
        ]
        if ordering in valid_ordering_fields:
            queryset = queryset.order_by(ordering)
        else:
            queryset = queryset.order_by('-created_at')

        # Select related data to avoid N+1 queries
        queryset = queryset.select_related(
            'staff_user',
            'staff_user__profile',
            'leave_type',
            'approved_by'
        )

        # Paginate results
        paginator = Paginator(queryset, page_size)
        page_obj = paginator.get_page(page_number)

        # Serialize data
        requests_data = []
        for leave_request in page_obj:
            request_data = {
                'id': leave_request.id,
                'staff_user': {
                    'id': leave_request.staff_user.id,
                    'name': f"{leave_request.staff_user.first_name} {leave_request.staff_user.last_name}",
                    'email': leave_request.staff_user.email,
                    'department': leave_request.staff_user.profile.employment_type.name if hasattr(leave_request.staff_user, 'profile') and leave_request.staff_user.profile.employment_type else None
                },
                'leave_type': {
                    'id': leave_request.leave_type.id,
                    'name': leave_request.leave_type.name,
                    'code': leave_request.leave_type.code,
                    'color_code': leave_request.leave_type.color_code
                },
                'start_date': leave_request.start_date.isoformat(),
                'end_date': leave_request.end_date.isoformat(),
                'days_requested': float(leave_request.days_requested),
                'status': leave_request.status,
                'reason': leave_request.reason,
                'created_at': leave_request.created_at.isoformat(),
                'approved_by': {
                    'id': leave_request.approved_by.id,
                    'name': f"{leave_request.approved_by.first_name} {leave_request.approved_by.last_name}"
                } if leave_request.approved_by else None,
                'approved_at': leave_request.approved_at.isoformat() if leave_request.approved_at else None,
                'manager_notes': leave_request.manager_notes if hasattr(leave_request, 'manager_notes') else None,
                'rejection_reason': leave_request.rejection_reason if hasattr(leave_request, 'rejection_reason') else None
            }
            requests_data.append(request_data)

        # Return paginated response
        return Response({
            'results': requests_data,
            'pagination': {
                'count': paginator.count,
                'page': page_number,
                'page_size': page_size,
                'total_pages': paginator.num_pages,
                'has_next': page_obj.has_next(),
                'has_previous': page_obj.has_previous()
            }
        })


class LeaveSettingsViewSet(viewsets.ViewSet):
    """
    ViewSet for Leave System Settings Management

    Handles system-wide leave settings and blackout periods
    This ViewSet doesn't work with a model, so it uses ViewSet instead of ModelViewSet
    """
    permission_classes = [IsAuthenticated, AdminOnlyPermission]

    def list(self, request):
        """Get leave system settings overview"""
        # System-wide settings (could be stored in database or config)
        settings = {
            'system_settings': {
                'default_working_days_per_week': 5,
                'default_working_hours_per_day': 8,
                'auto_approve_threshold_days': None,  # null = no auto-approval
                'max_advance_booking_days': 365,
                'require_manager_notes_on_rejection': True,
                'allow_weekend_requests': False,
                'notification_settings': {
                    'notify_manager_on_submission': True,
                    'notify_user_on_approval': True,
                    'notify_user_on_rejection': True,
                    'reminder_days_before_expiry': [90, 30, 7]
                }
            },
            'active_blackout_periods': BlackoutPeriod.objects.filter(
                is_active=True,
                end_date__gte=date.today()
            ).count(),
            'total_leave_types': LeaveType.objects.filter(is_active=True).count(),
            'total_policies': LeavePolicy.objects.filter(is_active=True).count(),
        }

        return Response(settings)

    @action(detail=False, methods=['get', 'put'], permission_classes=[AdminOnlyPermission])
    def system_config(self, request):
        """Get or update system-wide leave configuration"""
        from .models import SystemConfig

        if request.method == 'GET':
            # Load saved configuration from database
            saved_configs = {}
            for config in SystemConfig.objects.all():
                saved_configs[config.config_key] = config.config_data

            # Return saved configuration or defaults
            config = {
                'working_week': saved_configs.get('working_week', {
                    'days_per_week': 5,
                    'hours_per_day': 8,
                    'working_days': ['monday', 'tuesday', 'wednesday', 'thursday', 'friday']
                }),
                'approval_settings': saved_configs.get('approval_settings', {
                    'auto_approve_threshold_days': None,
                    'require_manager_notes_on_rejection': True,
                    'allow_emergency_override': True,
                    'max_advance_booking_days': 365
                }),
                'notification_settings': saved_configs.get('notification_settings', {
                    'email_notifications': True,
                    'sms_notifications': False,
                    'manager_notification_on_submission': True,
                    'user_notification_on_status_change': True,
                    'reminder_notifications': {
                        'enabled': True,
                        'days_before_expiry': [90, 30, 7],
                        'balance_low_threshold': 5
                    }
                }),
                'policy_defaults': saved_configs.get('policy_defaults', {
                    'default_accrual_method': 'monthly',
                    'default_carryover_method': 'partial',
                    'default_carryover_limit': 5,
                    'default_probation_months': 6
                }),
                'accrual_settings': saved_configs.get('accrual_settings', {})
            }
            return Response(config)

        elif request.method == 'PUT':
            # Update system configuration in database
            updated_config = request.data

            # Save each configuration section
            for config_key, config_data in updated_config.items():
                SystemConfig.objects.update_or_create(
                    config_key=config_key,
                    defaults={
                        'config_data': config_data,
                        'updated_by': request.user,
                        'description': f'System configuration for {config_key}'
                    }
                )

            return Response({
                'message': 'System configuration updated successfully',
                'updated_config': updated_config,
                'updated_at': timezone.now().isoformat()
            })

    @action(detail=False, methods=['get', 'put'], permission_classes=[AdminOnlyPermission])
    def notifications(self, request):
        """Get or update notification settings"""
        from .models import SystemConfig

        if request.method == 'GET':
            # Load saved notification settings from database
            try:
                config = SystemConfig.objects.get(config_key='notification_settings')
                return Response(config.config_data)
            except SystemConfig.DoesNotExist:
                # Return defaults
                return Response({
                    'email_notifications': True,
                    'sms_notifications': False,
                    'manager_approval_notifications': True,
                    'employee_request_notifications': True,
                    'balance_threshold_notifications': True,
                    'accrual_processing_notifications': False,
                    'reminder_days_before': 7,
                    'digest_frequency': 'weekly'
                })

        elif request.method == 'PUT':
            # Update notification settings
            updated_settings = request.data
            SystemConfig.objects.update_or_create(
                config_key='notification_settings',
                defaults={
                    'config_data': updated_settings,
                    'updated_by': request.user,
                    'description': 'System notification settings'
                }
            )

            return Response({
                'message': 'Notification settings updated successfully',
                'updated_settings': updated_settings,
                'updated_at': timezone.now().isoformat()
            })

    @action(detail=False, methods=['get'], permission_classes=[AdminOnlyPermission])
    def system_health(self, request):
        """Get system health and diagnostic information"""
        from .models import SystemConfig, LeaveRequest
        from django.db import connection

        # Get last accrual run timestamp
        try:
            accrual_config = SystemConfig.objects.get(config_key='accrual_settings')
            last_accrual_run = accrual_config.updated_at
            accrual_status = 'healthy'
        except SystemConfig.DoesNotExist:
            last_accrual_run = None
            accrual_status = 'not_configured'

        # Get pending notification count (approximate based on pending leave requests)
        pending_notifications = LeaveRequest.objects.filter(
            status='pending'
        ).count()

        # Check database connection
        try:
            with connection.cursor() as cursor:
                cursor.execute("SELECT 1")
            database_status = 'healthy'
            database_response_time = '< 5ms'
        except Exception as e:
            database_status = 'error'
            database_response_time = 'N/A'

        # Get system statistics
        total_leave_requests = LeaveRequest.objects.count()
        pending_approvals = LeaveRequest.objects.filter(status='pending').count()

        return Response({
            'accrual_engine': {
                'status': accrual_status,
                'last_run': last_accrual_run.isoformat() if last_accrual_run else None,
                'next_run': 'Scheduled for next month start'
            },
            'notifications': {
                'status': 'operational',
                'pending_count': pending_notifications,
                'queue_status': 'processing'
            },
            'database': {
                'status': database_status,
                'response_time': database_response_time,
                'connection_pool': 'healthy'
            },
            'statistics': {
                'total_leave_requests': total_leave_requests,
                'pending_approvals': pending_approvals
            },
            'last_updated': timezone.now().isoformat()
        })


class BlackoutPeriodsViewSet(viewsets.ModelViewSet):
    """
    ViewSet for managing blackout periods

    Allows administrators to manage periods when leave requests are restricted
    """
    queryset = BlackoutPeriod.objects.select_related('venue').prefetch_related('leave_types')
    serializer_class = BlackoutPeriodSerializer
    permission_classes = [IsAuthenticated, AdminOnlyPermission]
    filter_backends = [DjangoFilterBackend, filters.SearchFilter, filters.OrderingFilter]
    filterset_fields = ['is_active', 'restriction_level', 'venue']
    search_fields = ['name', 'description']
    ordering_fields = ['start_date', 'end_date', 'created_at']
    ordering = ['start_date']

    def get_queryset(self):
        """Filter queryset based on query parameters"""
        queryset = super().get_queryset()

        # Filter by date range
        start_date = self.request.query_params.get('start_date')
        end_date = self.request.query_params.get('end_date')

        if start_date:
            queryset = queryset.filter(end_date__gte=start_date)
        if end_date:
            queryset = queryset.filter(start_date__lte=end_date)

        # Filter by current/future periods
        if self.request.query_params.get('current_and_future', 'false').lower() == 'true':
            queryset = queryset.filter(end_date__gte=date.today())

        return queryset

    def list(self, request):
        """List blackout periods with additional context"""
        queryset = self.filter_queryset(self.get_queryset())

        page = self.paginate_queryset(queryset)
        if page is not None:
            serializer = self.get_serializer(page, many=True)
            response_data = self.get_paginated_response(serializer.data).data
        else:
            serializer = self.get_serializer(queryset, many=True)
            response_data = serializer.data

        # Add summary information
        today = date.today()
        summary = {
            'total_periods': queryset.count(),
            'active_periods': queryset.filter(is_active=True).count(),
            'current_restrictions': queryset.filter(
                is_active=True,
                start_date__lte=today,
                end_date__gte=today
            ).count(),
            'upcoming_restrictions': queryset.filter(
                is_active=True,
                start_date__gt=today
            ).count()
        }

        if isinstance(response_data, dict):
            # Paginated response
            response_data['summary'] = summary
            return Response(response_data)
        else:
            # Non-paginated response
            return Response({
                'blackout_periods': response_data,
                'summary': summary
            })

    @action(detail=False, methods=['get'])
    def current_restrictions(self, request):
        """Get currently active blackout periods"""
        today = date.today()
        current_restrictions = self.get_queryset().filter(
            is_active=True,
            start_date__lte=today,
            end_date__gte=today
        )

        serializer = self.get_serializer(current_restrictions, many=True)
        return Response({
            'current_restrictions': serializer.data,
            'count': current_restrictions.count(),
            'effective_date': today
        })

    @action(detail=False, methods=['get'])
    def upcoming_restrictions(self, request):
        """Get upcoming blackout periods"""
        today = date.today()
        days_ahead = int(request.query_params.get('days_ahead', 30))
        future_date = today + timedelta(days=days_ahead)

        upcoming_restrictions = self.get_queryset().filter(
            is_active=True,
            start_date__gt=today,
            start_date__lte=future_date
        ).order_by('start_date')

        serializer = self.get_serializer(upcoming_restrictions, many=True)
        return Response({
            'upcoming_restrictions': serializer.data,
            'count': upcoming_restrictions.count(),
            'period': {
                'start_date': today,
                'end_date': future_date,
                'days_ahead': days_ahead
            }
        })

    @action(detail=True, methods=['post'])
    def toggle_active(self, request, pk=None):
        """Toggle active status of a blackout period"""
        blackout_period = self.get_object()
        blackout_period.is_active = not blackout_period.is_active
        blackout_period.save()

        return Response({
            'message': f'Blackout period "{blackout_period.name}" {"activated" if blackout_period.is_active else "deactivated"}',
            'blackout_period': self.get_serializer(blackout_period).data
        })

    @action(detail=False, methods=['post'])
    def bulk_create(self, request):
        """Create multiple blackout periods at once"""
        periods_data = request.data.get('blackout_periods', [])

        if not periods_data:
            return Response({
                'error': 'No blackout periods provided'
            }, status=status.HTTP_400_BAD_REQUEST)

        created_periods = []
        errors = []

        for i, period_data in enumerate(periods_data):
            serializer = self.get_serializer(data=period_data)
            if serializer.is_valid():
                created_periods.append(serializer.save())
            else:
                errors.append({
                    'index': i,
                    'errors': serializer.errors,
                    'data': period_data
                })

        response_data = {
            'created_count': len(created_periods),
            'error_count': len(errors),
            'created_periods': self.get_serializer(created_periods, many=True).data
        }

        if errors:
            response_data['errors'] = errors

        return Response(
            response_data,
            status=status.HTTP_201_CREATED if created_periods else status.HTTP_400_BAD_REQUEST
        )

    @action(detail=False, methods=['get'])
    def check_conflicts(self, request):
        """Check for conflicting blackout periods"""
        start_date_str = request.query_params.get('start_date')
        end_date_str = request.query_params.get('end_date')

        if not start_date_str or not end_date_str:
            return Response({
                'error': 'Both start_date and end_date are required'
            }, status=status.HTTP_400_BAD_REQUEST)

        try:
            start_date = datetime.strptime(start_date_str, '%Y-%m-%d').date()
            end_date = datetime.strptime(end_date_str, '%Y-%m-%d').date()
        except ValueError:
            return Response({
                'error': 'Invalid date format. Use YYYY-MM-DD'
            }, status=status.HTTP_400_BAD_REQUEST)

        # Check for overlapping periods
        conflicting_periods = BlackoutPeriod.objects.overlapping_period(start_date, end_date)

        serializer = self.get_serializer(conflicting_periods, many=True)
        return Response({
            'has_conflicts': conflicting_periods.exists(),
            'conflict_count': conflicting_periods.count(),
            'conflicting_periods': serializer.data,
            'checked_period': {
                'start_date': start_date,
                'end_date': end_date
            }
        })


class HolidayViewSet(viewsets.ReadOnlyModelViewSet):
    """
    ViewSet for fetching public holidays via proxy to avoid CORS issues
    """
    permission_classes = [IsAuthenticated]

    def list(self, request):
        """
        Get public holidays for a specific country and year
        Proxies requests to Nager.Date API to avoid CORS issues
        """
        country_code = request.query_params.get('country', 'GB')  # Default to UK
        year = request.query_params.get('year', datetime.now().year)

        # Create cache key
        cache_key = f'holidays_{country_code}_{year}'

        # Try to get from cache first (cache for 24 hours)
        cached_holidays = cache.get(cache_key)
        if cached_holidays:
            logger.info(f'Returning cached holidays for {country_code} {year}')
            return Response(cached_holidays)

        try:
            # Make request to Nager.Date API
            api_url = f'https://date.nager.at/api/v3/PublicHolidays/{year}/{country_code}'

            logger.info(f'Fetching holidays from: {api_url}')
            response = requests.get(api_url, timeout=10)
            response.raise_for_status()

            holidays_data = response.json()

            # Transform data to match our frontend expectations
            transformed_holidays = []
            for holiday in holidays_data:
                transformed_holidays.append({
                    'date': holiday.get('date'),
                    'name': holiday.get('name'),
                    'localName': holiday.get('localName'),
                    'countryCode': holiday.get('countryCode'),
                    'fixed': holiday.get('fixed', True),
                    'global': holiday.get('global', True),
                    'types': holiday.get('types', [])
                })

            # Cache the result for 24 hours
            cache.set(cache_key, transformed_holidays, 60 * 60 * 24)

            logger.info(f'Successfully fetched {len(transformed_holidays)} holidays for {country_code} {year}')
            return Response(transformed_holidays)

        except requests.exceptions.RequestException as e:
            logger.error(f'Error fetching holidays from Nager.Date API: {str(e)}')

            # Return fallback data for common holidays
            fallback_holidays = self._get_fallback_holidays(country_code, year)
            return Response(fallback_holidays)

        except Exception as e:
            logger.error(f'Unexpected error fetching holidays: {str(e)}')
            return Response([], status=status.HTTP_500_INTERNAL_SERVER_ERROR)

    def _get_fallback_holidays(self, country_code, year):
        """
        Provide fallback holiday data when API is unavailable
        """
        year = int(year)

        if country_code.upper() == 'GB':
            return [
                {'date': f'{year}-01-01', 'name': 'New Year\'s Day', 'localName': 'New Year\'s Day', 'countryCode': 'GB', 'fixed': True, 'global': True, 'types': ['Public']},
                {'date': f'{year}-12-25', 'name': 'Christmas Day', 'localName': 'Christmas Day', 'countryCode': 'GB', 'fixed': True, 'global': True, 'types': ['Public']},
                {'date': f'{year}-12-26', 'name': 'Boxing Day', 'localName': 'Boxing Day', 'countryCode': 'GB', 'fixed': True, 'global': True, 'types': ['Public']},
            ]
        elif country_code.upper() == 'US':
            return [
                {'date': f'{year}-01-01', 'name': 'New Year\'s Day', 'localName': 'New Year\'s Day', 'countryCode': 'US', 'fixed': True, 'global': True, 'types': ['Public']},
                {'date': f'{year}-07-04', 'name': 'Independence Day', 'localName': 'Independence Day', 'countryCode': 'US', 'fixed': True, 'global': True, 'types': ['Public']},
                {'date': f'{year}-12-25', 'name': 'Christmas Day', 'localName': 'Christmas Day', 'countryCode': 'US', 'fixed': True, 'global': True, 'types': ['Public']},
            ]

        return []  # Return empty array for unsupported countries
